# -*- coding: utf-8 -*-
"""Módulo PFD Flowsheet — Diagrama de processo industrial."""

import sys
import math
import random
from PyQt5.QtWidgets import (
    QApplication,
    QMainWindow,
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QPushButton,
    QMessageBox,
    QGraphicsView,
    QGraphicsScene,
    QGraphicsItem,
    QListWidget,
    QListWidgetItem,
    QSplitter,
    QGraphicsPathItem,
    QMenu,
    QListView,
    QLineEdit,
    QLabel,
    QStackedWidget,
    QTextEdit,
    QTableWidget,
    QTableWidgetItem,
    QHeaderView,
    QGraphicsRectItem,
    QInputDialog,
    QFileDialog,
    QSizePolicy,
    QGraphicsEllipseItem,
    QToolBox,
    QGraphicsSimpleTextItem,
    QGraphicsTextItem,
    QDialog,
    QComboBox,
    QFormLayout,
    QDialogButtonBox,
)
from PyQt5.QtGui import (
    QPen,
    QBrush,
    QColor,
    QPainter,
    QPalette,
    QCursor,
    QPolygonF,
    QFont,
    QFontMetrics,
    QIcon,
    QPixmap,
    QPainterPath,
    QDrag,
    QLinearGradient,
)
from PyQt5.QtCore import (
    Qt,
    QRectF,
    QPointF,
    QMimeData,
    QByteArray,
    QDataStream,
    QIODevice,
    QSize,
    QPoint,
    QTimer,
    pyqtSignal,
    QObject,
    QSizeF,
)

from proeng.core.themes import T, THEMES, _ACTIVE
from proeng.core.utils import (
    _export_view,
    C_BG_APP,
    C_BG_NODE,
    C_BORDER,
    C_TEXT_MAIN,
    C_TEXT,
    C_PLACEHOLDER,
    C_BTN_ADD,
    C_BTN_DEL,
    C_LINE,
    C_BTN_SIB,
    C_BG_ROOT,
    C_BORDER_ROOT,
    W5H2_TYPES,
)
from proeng.core.toolbar import _make_toolbar, _hide_inner_toolbar
from proeng.core.base_module import BaseModule

import numpy as np
from collections import defaultdict

INDUSTRIAL_COMPONENTS = [
    "Água",
    "Vapor",
    "Vapor d'água",
    "Condensado",
    "Água de resfriamento",
    "NaOH",
    "KOH",
    "Ca(OH)₂",
    "Mg(OH)₂",
    "Al(OH)₃",
    "NH₃",
    "NH₄OH",
    "HCl",
    "H₂SO₄",
    "HNO₃",
    "H₃PO₄",
    "HF",
    "HBr",
    "HI",
    "H₃PO₄",
    "Ácido acético",
    "CH₃OH",
    "Etanol",
    "Propanol",
    "Butanol",
    "Isobutanol",
    "Glicerol",
    "Etileno glicol",
    "Metanol",
    "Acetona",
    "Benzeno",
    "Tolueno",
    "Xileno",
    "Etilbenzeno",
    "Cumeno",
    "Etileno",
    "Propileno",
    "Butadieno",
    "Isopreno",
    "Acetileno",
    "Buteno",
    "NaCl",
    "Na₂SO₄",
    "NaNO₃",
    "Na₂CO₃",
    "NaHCO₃",
    "KCl",
    "K₂SO₄",
    "KNO₃",
    "CaCO₃",
    "CaO",
    "CaSO₄",
    "CaCl₂",
    "MgO",
    "MgSO₄",
    "MgCl₂",
    "FeCl₃",
    "Fe₂O₃",
    "Fe₃O₄",
    "CuSO₄",
    "CuO",
    "CuCl₂",
    "ZnO",
    "ZnCl₂",
    "Al₂O₃",
    "AlCl₃",
    "SiO₂",
    "TiO₂",
    "C (Carvão)",
    "Coque",
    "Grafite",
    "CO",
    "CO₂",
    "CH₄",
    "C₂H₆",
    "C₃H₈",
    "C₄H₁₀",
    "H₂",
    "O₂",
    "N₂",
    "Ar",
    "He",
    "SO₂",
    "SO₃",
    "H₂S",
    "Cl₂",
    "Br₂",
    "I₂",
    "Ozone",
    "Urea",
    "NH₄NO₃",
    "(NH₄)₂SO₄",
    "MAP",
    "DAP",
    "NPK",
    "Superfosfato",
    "Glucose",
    "Sacarose",
    "Amido",
    "Celulose",
    "Proteína",
    "Lipídio",
    "Lactose",
    "Formaldeído",
    "Formol",
    "Acetaldeído",
    "Ácido fórmico",
    "Ácido lático",
    "Fenol",
    "Cresol",
    "Naftaleno",
    "Antraceno",
    "Fenantreno",
    "PVC",
    "PE",
    "PP",
    "PS",
    "PET",
    "PA",
    "PMMA",
    "Teflon",
    "Metano",
    "Etano",
    "Propano",
    "Butano",
    "Pentano",
    "Hexano",
    "Heptano",
    "Octano",
    "Dietanolamina",
    "Monoetanolamina",
    "Trietanolamina",
    "Piperazina",
    "Hidrazina",
    "Hidróxido de amônio",
    "Carbonato de amônio",
    "Bicarbonato de amônio",
    "Sulfato de amônio",
    "Fosfato de amônio",
    "Cloreto de amônio",
    "Nitrato de amônio",
    "Cimento",
    "Gesso",
    "Cal",
    "Dolomita",
    "Magnesita",
    "Bauxita",
    "Areia",
    "Brita",
    "Cascalho",
    "Argila",
    "Silte",
    "Minério de ferro",
    "Minério de cobre",
    "Minério de zinco",
    "Minério de chumbo",
    "Bauxita",
    "Lama vermelha",
    "Lodo",
    "Resíduo",
    "Efluente",
    "Água residual",
    "Óleo",
    "Óleo cru",
    "Óleo diesel",
    "Gasolina",
    "Querosene",
    "Asfalteno",
    "Betume",
    "Parafina",
    "Cera",
    "Resina",
    "Alcatrão",
    "Piche",
    "Lixívia",
    "Leite de cal",
    "Borra",
    "Filtrado",
    "Cake",
    "Concentrado",
    "Produtos de reação",
    "Produtos intermediários",
    "Produtos principais",
    "Subprodutos",
    "Impurezas",
    "Catalisador",
    "Inibidor",
    "Promotor",
    "Aditivo",
    "Corante",
    "Floculante",
    "Coagulante",
    "Dispersante",
    "Surfactante",
    "Detergente",
    "Sabão",
    "Emulsificante",
    "Espumante",
]

from PyQt5.QtWidgets import QDialog, QComboBox, QFormLayout, QDialogButtonBox


class TerminalConfigDialog(QDialog):
    """Diálogo industrial para escolher Tipo e Nome do fluxo."""

    def __init__(self, parent=None, suggested_type="Saída"):
        super().__init__(parent)
        self.setWindowTitle("Configurar Terminal")
        self.setMinimumWidth(320)
        t = T()

        layout = QVBoxLayout(self)
        form = QFormLayout()

        self.type_combo = QComboBox()
        self.type_combo.addItems(["Entrada", "Saída"])
        self.type_combo.setCurrentText(suggested_type)

        self.name_edit = QLineEdit()
        self.name_edit.setPlaceholderText("Ex: Água, Vapor, MP-01...")
        self.name_edit.setFocus()

        form.addRow("Tipo de Fluxo:", self.type_combo)
        form.addRow("Nome do Fluxo:", self.name_edit)
        layout.addLayout(form)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

        self.setStyleSheet(f"""
            QDialog {{ background-color: {t["bg_app"]}; color: {t["text"]}; }}
            QLabel {{ color: {t["text"]}; font-weight: bold; }}
            QComboBox, QLineEdit {{ 
                background: {t["bg_card"]}; 
                border: 1px solid {t["accent"]}; 
                border-radius: 4px; 
                padding: 6px; 
                color: {t["text"]}; 
            }}
            QPushButton {{ 
                background: {t["bg_card"]}; 
                color: {t["text"]}; 
                border-radius: 4px; 
                padding: 6px 15px; 
                min-width: 80px;
            }}
            QPushButton:hover {{ background: {t["accent"]}; color: white; }}
        """)

    def get_values(self):
        return self.type_combo.currentText(), self.name_edit.text()


class StreamEditorDialog(QDialog):
    """Diálogo avançado para inserir dados de balanço (Vazão Total e Porcentagens)."""

    def __init__(self, edge, parent=None, global_components=None):
        super().__init__(parent)
        self.edge = edge
        self.global_components = global_components or INDUSTRIAL_COMPONENTS
        self.setWindowTitle(f"Dados da Corrente: {edge.pipe_name}")
        self.setMinimumWidth(600)
        self.setMinimumHeight(450)
        t = T()

        layout = QVBoxLayout(self)

        # Header: Vazão Total
        header_layout = QFormLayout()
        self.total_flow_edit = QLineEdit()
        current_total = sum(edge.flow_data.values())
        self.total_flow_edit.setText(f"{current_total:.2f}")
        self.total_flow_edit.textChanged.connect(self.sync_all_from_total)
        header_layout.addRow("Vazão Mássica TOTAL (kg/h):", self.total_flow_edit)
        layout.addLayout(header_layout)

        # Tabela: Comp, %, Flow
        self.table = QTableWidget(0, 3)
        self.table.setHorizontalHeaderLabels(
            ["Componente", "Porcentagem (%)", "Vazão (kg/h)"]
        )
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.itemChanged.connect(self.on_item_changed)
        layout.addWidget(self.table)

        # Info de Soma
        self.sum_label = QLabel("Soma das Porcentagens: 0.00%")
        self.sum_label.setStyleSheet("font-weight: bold; padding: 5px;")
        layout.addWidget(self.sum_label)

        # Popula a tabela
        self._updating = True  # Flag para evitar recursão infinita no sync
        for comp, flow in self.edge.flow_data.items():
            perc = (flow / current_total * 100) if current_total > 0 else 0.0
            self.add_row(comp, perc, flow)
        self._updating = False
        self.update_sum_label()

        btn_layout = QHBoxLayout()
        btn_add = QPushButton("➕ Adicionar Componente")
        btn_add.clicked.connect(lambda: self.add_row("", 0.0, 0.0))
        btn_layout.addWidget(btn_add)

        btn_common = QPushButton("📋 Comuns (10)")
        btn_common.setToolTip("Adiciona 10 componentes comuns da indústria química")
        btn_common.clicked.connect(self.add_common_components)
        btn_layout.addWidget(btn_common)
        layout.addLayout(btn_layout)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.save_data)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

        self.setStyleSheet(f"""
            QDialog {{ background-color: {t["bg_app"]}; color: {t["text"]}; font-family: 'Segoe UI'; }}
            QTableWidget {{ background-color: {t["bg_card"]}; color: {t["text"]}; border: 1px solid {t["accent"]}; }}
            QLineEdit {{ background: {t["bg_card"]}; color: {t["text"]}; border: 1px solid {t["accent"]}; padding: 5px; }}
            QHeaderView::section {{ background-color: {t["toolbar_bg"]}; color: {t["text"]}; font-weight: bold; }}
            QLabel {{ color: {t["text"]}; }}
        """)

    def add_row(self, comp="", perc=0.0, flow=0.0):
        self._updating = True
        row = self.table.rowCount()
        self.table.insertRow(row)

        combo = QComboBox()
        combo.setEditable(True)
        combo.addItems(self.global_components)
        if comp:
            combo.setCurrentText(comp)
        self.table.setCellWidget(row, 0, combo)

        self.table.setItem(row, 1, QTableWidgetItem(f"{perc:.2f}"))
        self.table.setItem(row, 2, QTableWidgetItem(f"{flow:.2f}"))
        self._updating = False

    def add_common_components(self):
        """Adiciona 10 componentes comuns da indústria química."""
        common = [
            "Água",
            "NaOH",
            "HCl",
            "H₂SO₄",
            "NaCl",
            "Metanol",
            "Etanol",
            "Vapor",
            "CO₂",
            "O₂",
        ]
        try:
            total = float(self.total_flow_edit.text().replace(",", "."))
        except:
            total = 1000.0
        flow_per_comp = total / len(common) / 10 if total > 0 else 100.0
        for comp in common:
            perc = 10.0
            self.add_row(comp, perc, flow_per_comp)
        self.update_sum_label()

    def on_item_changed(self, item):
        if self._updating:
            return
        self._updating = True
        row, col = item.row(), item.column()
        try:
            total = float(self.total_flow_edit.text().replace(",", "."))
            if col == 1:  # Alterou % -> Atualiza Vazão
                perc = float(item.text().replace(",", "."))
                flow = total * (perc / 100.0)
                self.table.item(row, 2).setText(f"{flow:.2f}")
            elif col == 2:  # Alterou Vazão -> Atualiza %
                flow = float(item.text().replace(",", "."))
                perc = (flow / total * 100.0) if total > 0 else 0.0
                self.table.item(row, 1).setText(f"{perc:.2f}")
        except:
            pass
        self._updating = False
        self.update_sum_label()

    def sync_all_from_total(self):
        if self._updating:
            return
        self._updating = True
        try:
            total = float(self.total_flow_edit.text().replace(",", "."))
            for r in range(self.table.rowCount()):
                perc_txt = self.table.item(r, 1).text().replace(",", ".")
                flow = total * (float(perc_txt) / 100.0)
                self.table.item(r, 2).setText(f"{flow:.2f}")
        except:
            pass
        self._updating = False

    def update_sum_label(self):
        total_p = 0.0
        for r in range(self.table.rowCount()):
            try:
                total_p += float(self.table.item(r, 1).text().replace(",", "."))
            except:
                pass
        self.sum_label.setText(f"Soma das Porcentagens: {total_p:.2f}%")
        color = "#ff4444" if abs(total_p - 100) > 0.01 and total_p > 0 else T()["text"]
        self.sum_label.setStyleSheet(f"color: {color}; font-weight: bold;")

    def save_data(self):
        new_data = {}
        for row in range(self.table.rowCount()):
            comp_widget = self.table.cellWidget(row, 0)
            flow_item = self.table.item(row, 2)
            comp = comp_widget.currentText().strip() if comp_widget else ""
            if comp and flow_item:
                try:
                    val = float(flow_item.text().replace(",", "."))
                    new_data[comp] = val
                except:
                    pass
        self.edge.flow_data = new_data
        self.edge.adjust()

        if hasattr(self, "global_components"):
            view = self.edge.scene().views()[0]
            parent_widget = getattr(view, "parentWidget", None)
            if parent_widget and hasattr(parent_widget, "global_components"):
                for comp in new_data:
                    if comp not in parent_widget.global_components:
                        parent_widget.global_components.append(comp)

        view = self.edge.scene().views()[0]
        if hasattr(view, "parentWidget") and hasattr(
            view.parentWidget(), "solve_mass_balance"
        ):
            view.parentWidget().solve_mass_balance()

        self.accept()


class EquipmentEditorDialog(QDialog):
    """Diálogo avançado para configurar o desempenho (Vazão Fixa vs Fração)."""

    def __init__(self, node, parent=None):
        super().__init__(parent)
        self.node = node
        self.setWindowTitle(f"Desempenho: {node.symbol_type}")
        self.setMinimumWidth(700)
        self.setMinimumHeight(500)
        t = T()

        layout = QVBoxLayout(self)
        layout.addWidget(
            QLabel(
                "Defina como o componente é separado nas saídas.\nUse 'Fator' (0.0-1.0) ou 'Vazão' (kg/h fixa)."
            )
        )

        self.out_edges = [e for e in node.edges if e.source_node == node]
        self.out_ports = sorted(list(set(e.source_port for e in self.out_edges)))

        self.components = set()
        for e in node.edges:
            if e.dest_node == node:
                self.components.update(e.flow_data.keys())
        self.components = sorted(list(self.components))

        if not self.out_ports and not self.components:
            layout.addWidget(
                QLabel("⚠️ Conecte pelo menos uma entrada ou saída primeiro.")
            )
        elif not self.out_ports:
            layout.addWidget(
                QLabel(
                    "⚠️ O equipamento precisa de saídas configuradas para definir a distribuição de componentes."
                )
            )
        elif not self.components:
            layout.addWidget(
                QLabel("⚠️ Adicione componentes nas correntes de entrada primeiro.")
            )
        else:
            # Tabela complexa: Componente | Porta | Tipo (Combo) | Valor
            self.table = QTableWidget(len(self.components) * len(self.out_ports), 4)
            self.table.setHorizontalHeaderLabels(
                ["Componente", "Porta Saída", "Modo", "Valor"]
            )
            self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

            row = 0
            for comp in self.components:
                # Recupera config atual: { comp: { port: {"mode": "perc", "val": 0.5} } }
                comp_cfg = node.split_config.get(comp, {})
                for port in self.out_ports:
                    cfg = comp_cfg.get(
                        port, {"mode": "perc", "val": 1.0 / len(self.out_ports)}
                    )

                    self.table.setItem(row, 0, QTableWidgetItem(comp))
                    self.table.setItem(row, 1, QTableWidgetItem(port))

                    combo = QComboBox()
                    combo.addItems(["Fração (%)", "Vazão Fixa (kg/h)"])
                    combo.setCurrentIndex(0 if cfg["mode"] == "perc" else 1)
                    self.table.setCellWidget(row, 2, combo)

                    val_item = QTableWidgetItem(str(cfg["val"]))
                    self.table.setItem(row, 3, val_item)
                    row += 1
            layout.addWidget(self.table)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.save_config)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

        self.setStyleSheet(
            f"QDialog{{background:{t['bg_app']};color:{t['text']};}} QTableWidget{{background:{t['bg_card']};color:{t['text']};}}"
        )

    def save_config(self):
        if not hasattr(self, "table"):
            self.accept()
            return
        new_config = {}
        for r in range(self.table.rowCount()):
            comp = self.table.item(r, 0).text()
            port = self.table.item(r, 1).text()
            mode_idx = self.table.cellWidget(r, 2).currentIndex()
            val_txt = self.table.item(r, 3).text().replace(",", ".")
            try:
                val = float(val_txt)
            except:
                val = 0.0

            if comp not in new_config:
                new_config[comp] = {}
            new_config[comp][port] = {
                "mode": "perc" if mode_idx == 0 else "fixed",
                "val": val,
            }

        self.node.split_config = new_config
        self.accept()


EQUIPMENT_ALIASES = {
    "Bomba Centrífuga": "Bomba",
    "Bomba Volumétrica": "Bomba_V",
    "Compressores": "Compressor",
    "Trocador Casco-Tubo": "Trocador",
    "Reator CSTR": "Reator",
    "Tanque Aberto": "Tanque",
    "Vaso Vertical": "Vaso",
    "Tanque Misturador": "Misturador",
    "Secador Rotativo": "Secador",
    "Válvula Globo": "Válvula",
    "PSV (Alívio)": "PSV",
    "Válvula de Bloqueio": "Válvula",
    "Forno": "Forno",
    "Fornalha": "Forno",
    "Lavagem": "Lavadora",
    "Peneira": "Peneira",
    "Filtro": "Filtro",
    "Separador Bifásico": "Separador_B",
    "Vaso Horizontal": "Vaso_H",
    "Esfera de Gás": "Esfera",
    "Silo": "Silo",
    "Compressor": "Compressor",
    "Soprador": "Soprador",
    "Exaustor": "Fan",
    "Turbina": "Turbina",
    "Ejetor": "Ejetor",
    "Trocador de Placas": "Trocador_P",
    "Caldeira": "Caldeira",
    "Torre de Resfriamento": "Resfriamento",
    "Evaporador": "Evaporador",
    "Aquecedor Elétrico": "Heat_E",
    "Reator Tubular (PFR)": "Reator_T",
    "Moinho": "Moinho",
    "Britador": "Britador",
    "Válvula Borboleta": "Válvula_B",
    "Válvula de Controle": "Válvula_C",
    "Válvula de Retenção": "Válvula_R",
    "Flare": "Flare",
    "Chaminé": "Chaminé",
    "Correia Transportadora": "Correia",
    "Rosca Transportadora": "Rosca",
    "Elevador de Canecas": "Elevador",
    "Refervedor Kettle": "Refervedor",
    "Condensador de Casco": "Trocador",
    "Extrusora": "Extrusora",
    "Leito Fluidizado": "Reator_F",
    "Precipitador Eletrostático": "ESP",
    "Filtro Rotativo": "Filtro_R",
    "Coluna de Recheio": "Torre_R",
    "Válvula Solenoide": "Válvula_S",
    "Tanque de Teto Flutuante": "Tanque_F",
    "LPG Esfera": "Esfera",
    "Britador de Mandíbula": "Britador",
    "Peneira Vibratória": "Peneira_V",
    "Transmissor de Pressão": "Inst_P",
    "Transmissor de Temperatura": "Inst_T",
    "Transmissor de Vazão": "Inst_V",
    "Transmissor de Nível": "Inst_N",
    "Analisador": "Inst_A",
    "Válvula de Segurança (SRV)": "PSV",
    "Secador de Spray": "Spray_D",
    "Misturador Estático": "Mist_E",
    "Filtro de Mangas": "Bag_F",
    "Reator de Batelada": "Reator_B",
    "Cristalizador": "Cristal",
    "Filtro Prensa": "Filtro_P",
    "Adsorvedor": "Vaso",
    "Agitador de Pá": "Misturador",
    "Válvula de Pé": "Válvula",
    "Transportador Pneumático": "Pneum_C",
    "Válvula de Macho": "Válvula",
    "Filtro Cartucho": "Filtro",
    "Bomba de Vácuo": "Bomba",
    "Ejetor de Vapor": "Ejetor",
    "Resfriador de Filme": "Trocador",
    "Peneira Rotativa": "Peneira",
    "Hidrociclone": "Ciclone",
    "Ciclone de Gases": "Ciclone",
    "Talha Elétrica": "Box",
    "Painel de Controle": "Box",
    "Válvula de Esfera": "Válvula_B",
}


def draw_equipment(painter, symbol_type, size, is_icon=False, theme=None):
    s = size / 2
    if theme is None:
        try:
            theme = T()
        except:
            theme = THEMES["dark"]

    pen_color = theme["text"] if is_icon else theme["accent"]
    _bg_node = theme["bg_card"]

    # Efeito "Glass Neon" Industrial Dinâmico
    grad = QLinearGradient(QPointF(-size, -size), QPointF(size, size))
    c_base = QColor(theme["accent"])

    # Opacidade ajustada para o tema (mais sutil no Light)
    alpha1 = 60 if theme["name"] == "dark" else 40
    alpha2 = 20 if theme["name"] == "dark" else 10

    c_fill1 = QColor(c_base.red(), c_base.green(), c_base.blue(), alpha1)
    c_fill2 = QColor(c_base.red(), c_base.green(), c_base.blue(), alpha2)
    grad.setColorAt(0, c_fill1)
    grad.setColorAt(1, c_fill2)

    pw = 1.2 if is_icon else 1.8
    default_pen = QPen(QColor(pen_color), pw, Qt.SolidLine, Qt.RoundCap, Qt.RoundJoin)
    painter.setPen(default_pen)
    painter.setBrush(QBrush(grad))

    st = EQUIPMENT_ALIASES.get(symbol_type, symbol_type)

    if st == "Vaso":
        painter.drawRoundedRect(QRectF(-s * 0.8, -s, s * 1.6, s * 2), s * 0.6, s * 0.6)
    elif st == "Vaso_H":
        painter.drawRoundedRect(
            QRectF(-s * 1.8, -s * 0.8, s * 3.6, s * 1.6), s * 0.6, s * 0.6
        )
    elif st == "Tanque":
        painter.drawRect(QRectF(-s * 0.8, -s * 0.5, s * 1.6, s * 1.5))
        painter.drawArc(QRectF(-s * 0.8, -s * 1.0, s * 1.6, s * 1.0), 0, 180 * 16)
    elif st == "Tanque Fechado":
        painter.drawRect(QRectF(-s * 0.8, -s, s * 1.6, s * 2))
        painter.drawArc(QRectF(-s * 0.8, -s * 1.2, s * 1.6, s * 0.8), 0, 180 * 16)
        painter.drawArc(QRectF(-s * 0.8, s * 0.6, s * 1.6, s * 0.8), 180 * 16, 180 * 16)
    elif st == "Esfera de Gás":
        painter.drawEllipse(QRectF(-s, -s, s * 2, s * 2))
        painter.drawLine(QPointF(-s * 0.6, s * 0.8), QPointF(-s * 0.8, s * 1.5))
        painter.drawLine(QPointF(s * 0.6, s * 0.8), QPointF(s * 0.8, s * 1.5))
    elif st == "Silo":
        painter.drawRect(QRectF(-s * 0.8, -s * 1.5, s * 1.6, s * 2))
        poly = QPolygonF(
            [QPointF(-s * 0.8, s * 0.5), QPointF(s * 0.8, s * 0.5), QPointF(0, s * 1.5)]
        )
        painter.drawPolygon(poly)
    elif st == "Separador_B":
        painter.drawRoundedRect(
            QRectF(-s * 1.5, -s * 0.6, s * 3.0, s * 1.2), s * 0.6, s * 0.6
        )
        painter.drawLine(QPointF(-s * 1.5, 0), QPointF(s * 1.5, 0))
    elif st == "Separador Trifásico":
        painter.drawRoundedRect(
            QRectF(-s * 1.5, -s * 0.6, s * 3.0, s * 1.2), s * 0.6, s * 0.6
        )
        painter.drawLine(QPointF(-s * 1.5, -s * 0.2), QPointF(s * 1.5, -s * 0.2))
        painter.drawLine(QPointF(-s * 1.5, s * 0.2), QPointF(s * 1.5, s * 0.2))
    elif st == "Bomba":
        painter.drawEllipse(QRectF(-s * 0.8, -s * 0.8, s * 1.6, s * 1.6))
        poly = QPolygonF(
            [
                QPointF(-s * 0.3, -s * 0.7),
                QPointF(s * 0.3, -s * 0.7),
                QPointF(0, -s * 1.4),
            ]
        )
        painter.drawPolygon(poly)
    elif st == "Bomba Volumétrica":
        painter.drawEllipse(QRectF(-s * 0.8, -s * 0.8, s * 1.6, s * 1.6))
        painter.drawEllipse(QRectF(-s * 0.4, -s * 0.4, s * 0.8, s * 0.8))
        painter.drawLine(QPointF(0, -s * 0.8), QPointF(0, s * 0.8))
    elif st == "Soprador":
        painter.drawEllipse(QRectF(-s * 0.8, -s * 0.6, s * 1.2, s * 1.2))
        painter.drawRect(QRectF(s * 0.2, -s * 0.6, s * 0.5, s * 0.5))
    elif st == "Exaustor":
        painter.drawEllipse(QRectF(-s * 0.8, -s * 0.8, s * 1.6, s * 1.6))
        painter.drawLine(QPointF(-s * 0.6, 0), QPointF(s * 0.6, 0))
        painter.drawLine(QPointF(0, -s * 0.6), QPointF(0, s * 0.6))
    elif st == "Turbina":
        poly = QPolygonF(
            [
                QPointF(s * 0.8, s * 0.8),
                QPointF(-s * 0.8, s * 0.4),
                QPointF(-s * 0.8, -s * 0.4),
                QPointF(s * 0.8, -s * 0.8),
            ]
        )
        painter.drawPolygon(poly)
    elif st == "Ejetor":
        poly = QPolygonF(
            [
                QPointF(-s, -s * 0.3),
                QPointF(-s * 0.2, -s * 0.1),
                QPointF(s, -s * 0.3),
                QPointF(s, s * 0.3),
                QPointF(-s * 0.2, s * 0.1),
                QPointF(-s, s * 0.3),
            ]
        )
        painter.drawPolygon(poly)
    elif st == "Trocador":
        painter.drawEllipse(QRectF(-s, -s, s * 2, s * 2))
        painter.drawLine(QPointF(-s, 0), QPointF(-s * 0.5, -s * 0.5))
        painter.drawLine(QPointF(-s * 0.5, -s * 0.5), QPointF(0, s * 0.5))
        painter.drawLine(QPointF(0, s * 0.5), QPointF(s * 0.5, -s * 0.5))
        painter.drawLine(QPointF(s * 0.5, -s * 0.5), QPointF(s, 0))
    elif st == "Trocador_P":
        painter.drawRect(QRectF(-s * 1.2, -s * 0.8, s * 2.4, s * 1.6))
        for i in [-0.8, -0.4, 0, 0.4, 0.8]:
            painter.drawLine(QPointF(s * i, -s * 0.8), QPointF(s * i, s * 0.8))
    elif st == "Permutador a Ar":
        painter.drawRect(QRectF(-s, -s * 0.5, s * 2, s))
        painter.drawLine(QPointF(-s * 0.5, -s * 0.5), QPointF(s * 0.5, -s * 1.2))
        painter.drawLine(QPointF(s * 0.5, -s * 0.5), QPointF(-s * 0.5, -s * 1.2))
        painter.drawEllipse(QRectF(-s * 0.6, -s * 1.3, s * 1.2, s * 0.8))
    elif st == "Resfriador de Topo":
        painter.drawEllipse(QRectF(-s * 0.8, -s * 0.8, s * 1.6, s * 1.6))
        painter.drawLine(QPointF(-s * 0.8, -s * 0.8), QPointF(s * 0.8, s * 0.8))
    elif st == "Forno":
        painter.drawRect(QRectF(-s, -s * 0.5, s * 2, s * 1.5))
        poly = QPolygonF(
            [QPointF(-s, -s * 0.5), QPointF(0, -s * 1.5), QPointF(s, -s * 0.5)]
        )
        painter.drawPolygon(poly)
        painter.drawLine(QPointF(-s * 0.6, s * 0.6), QPointF(s * 0.6, s * 0.2))
        painter.drawLine(QPointF(s * 0.6, s * 0.2), QPointF(-s * 0.6, -0.2))
    elif st == "Caldeira":
        painter.drawRoundedRect(
            QRectF(-s * 1.5, -s * 0.8, s * 3.0, s * 1.6), s * 0.4, s * 0.4
        )
        for i in [-0.4, 0, 0.4]:
            painter.drawLine(QPointF(-s * 1.5, s * i), QPointF(s * 1.5, s * i))
    elif st == "Torre de Resfriamento":
        poly = QPolygonF(
            [
                QPointF(-s, s * 1.5),
                QPointF(s, s * 1.5),
                QPointF(s * 0.6, -s * 1.2),
                QPointF(-s * 0.6, -s * 1.2),
            ]
        )
        painter.drawPolygon(poly)
        painter.drawLine(QPointF(-s * 0.4, -s * 1.4), QPointF(s * 0.4, -s * 1.4))
        painter.drawLine(QPointF(0, -s * 1.2), QPointF(0, -s * 1.4))
    elif st == "Evaporador":
        painter.drawRoundedRect(
            QRectF(-s * 0.6, -s * 1.5, s * 1.2, s * 2), s * 0.6, s * 0.6
        )
        painter.drawRect(QRectF(-s * 0.6, s * 0.5, s * 1.2, s * 1.0))
    elif st == "Heat_E":
        painter.drawRect(QRectF(-s * 0.6, -s, s * 1.2, s * 2))
        poly = QPolygonF(
            [
                QPointF(-s * 0.4, 0),
                QPointF(0, -s * 0.6),
                QPointF(s * 0.4, 0),
                QPointF(0, s * 0.6),
            ]
        )
        painter.drawPolygon(poly)
    elif st.startswith("Inst_"):
        painter.drawEllipse(QRectF(-s * 0.8, -s * 0.8, s * 1.6, s * 1.6))
        txt = st.split("_")[1]
        painter.drawText(
            QRectF(-s * 0.8, -s * 0.8, s * 1.6, s * 1.6), Qt.AlignCenter, txt
        )
    elif st == "Spray_D":
        painter.drawRect(QRectF(-s, -s * 1.2, s * 2, s * 1.2))
        poly = QPolygonF([QPointF(-s, 0), QPointF(0, s * 1.5), QPointF(s, 0)])
        painter.drawPolygon(poly)
    elif st == "Mist_E":
        painter.drawRect(QRectF(-s * 2, -s * 0.3, s * 4, s * 0.6))
        for i in range(5):
            painter.drawLine(
                QPointF(-s * 2 + i * s * 0.8, -s * 0.3),
                QPointF(-s * 1.6 + i * s * 0.8, s * 0.3),
            )
    elif st == "Bag_F":
        painter.drawRect(QRectF(-s, -s, s * 2, s * 1.5))
        for i in range(3):
            painter.drawRect(QRectF(-s * 0.8 + i * s * 0.6, -s * 0.8, s * 0.4, s * 2))
    elif st == "Reator_B":
        painter.drawRoundedRect(QRectF(-s, -s, s * 2, s * 2.5), s * 0.4, s * 0.4)
        painter.drawLine(QPointF(0, -s * 0.5), QPointF(0, s * 0.5))
        painter.drawLine(QPointF(-s * 0.4, s * 0.5), QPointF(s * 0.4, s * 0.5))
    elif st == "Cristal":
        painter.drawRoundedRect(QRectF(-s, -s * 1.5, s * 2, s * 3), s * 0.5, s * 0.5)
        for i in range(10):
            painter.drawPoint(
                QPointF(random.uniform(-s * 0.5, s * 0.5), random.uniform(-s, s))
            )
    elif st == "Filtro_P":
        painter.drawRect(QRectF(-s * 1.5, -s, s * 3, s * 2))
        for i in range(6):
            painter.drawLine(
                QPointF(-s * 1.5 + i * s * 0.5, -s), QPointF(-s * 1.5 + i * s * 0.5, s)
            )
    elif st == "Pneum_C":
        painter.drawRect(QRectF(-s * 2, -s * 0.2, s * 4, s * 0.4))
        painter.drawEllipse(QRectF(s * 1.5, -s * 0.5, s, s))
    elif st == "Vaso_H":
        painter.drawRoundedRect(
            QRectF(-s * 1.8, -s * 0.8, s * 3.6, s * 1.6), s * 0.6, s * 0.6
        )
        painter.drawLine(QPointF(-s * 1.2, s * 0.8), QPointF(-s * 1.2, s * 1.1))
        painter.drawLine(QPointF(s * 1.2, s * 0.8), QPointF(s * 1.2, s * 1.1))
    elif st == "Compressor":
        painter.drawEllipse(QRectF(-s, -s, s * 2, s * 2))
        painter.drawEllipse(QRectF(-s * 0.6, -s * 0.6, s * 1.2, s * 1.2))
        painter.drawLine(QPointF(0, -s), QPointF(0, s))
    elif st == "Fan":
        painter.drawEllipse(QRectF(-s, -s, s * 2, s * 2))
        for i in range(4):
            ang = i * 90
            painter.drawLine(
                QPointF(0, 0),
                QPointF(
                    s * math.cos(math.radians(ang)), s * math.sin(math.radians(ang))
                ),
            )
    elif st == "Turbina":
        poly = QPolygonF(
            [QPointF(-s, -s * 0.6), QPointF(s, -s), QPointF(s, s), QPointF(-s, s * 0.6)]
        )
        painter.drawPolygon(poly)
    elif st == "Ejetor":
        poly = QPolygonF(
            [
                QPointF(-s, -s * 0.3),
                QPointF(-s * 0.2, -s * 0.2),
                QPointF(s * 0.2, -s * 0.4),
                QPointF(s, -s * 0.4),
                QPointF(s, s * 0.4),
                QPointF(s * 0.2, s * 0.4),
                QPointF(-s * 0.2, s * 0.2),
                QPointF(-s, s * 0.3),
            ]
        )
        painter.drawPolygon(poly)
    elif st == "Trocador_P":
        painter.drawRect(QRectF(-s * 0.8, -s * 1.2, s * 1.6, s * 2.4))
        for i in range(5):
            painter.drawLine(
                QPointF(-s * 0.8, -s + i * s * 0.5),
                QPointF(s * 0.8, -s + i * s * 0.5 + 0.2),
            )
    elif st == "Resfriamento":
        painter.drawRect(QRectF(-s, -s, s * 2, s * 2))
        painter.drawLine(QPointF(-s, -s * 0.7), QPointF(s, -s * 0.7))
        painter.drawEllipse(QRectF(-s * 0.4, -s * 1.2, s * 0.8, s * 0.4))
    elif st == "Evaporador":
        painter.drawRoundedRect(
            QRectF(-s * 0.7, -s * 1.5, s * 1.4, s * 3), s * 0.5, s * 0.5
        )
        for i in range(4):
            painter.drawArc(
                QRectF(-s * 0.4, -s + i * s * 0.5, s * 0.8, s * 0.4), 0, 180 * 16
            )
    elif st == "Reator_T":
        painter.drawRect(QRectF(-s * 1.6, -s * 0.5, s * 3.2, s))
        for i in [-1.0, -0.5, 0, 0.5, 1.0]:
            painter.drawEllipse(QRectF(s * i - s * 0.2, -s * 0.4, s * 0.4, s * 0.8))
    elif st == "Moinho":
        painter.drawEllipse(QRectF(-s, -s, s * 2, s * 2))
        painter.drawPoint(QPointF(0, 0))
        painter.drawEllipse(QRectF(-s * 0.5, -s * 0.5, s, s))
    elif st == "Britador":
        painter.drawLine(QPointF(-s, -s), QPointF(0, s))
        painter.drawLine(QPointF(s, -s), QPointF(0, s))
        painter.drawLine(QPointF(-s * 0.5, -s), QPointF(s * 0.5, -s))
    elif st == "Válvula_B":
        poly = QPolygonF(
            [
                QPointF(-s, -s * 0.6),
                QPointF(s, s * 0.6),
                QPointF(s, -s * 0.6),
                QPointF(-s, s * 0.6),
            ]
        )
        painter.drawPolygon(poly)
        painter.drawEllipse(QRectF(-s * 0.2, -s * 0.2, s * 0.4, s * 0.4))
    elif st == "Válvula_C":
        poly = QPolygonF(
            [
                QPointF(-s, -s * 0.6),
                QPointF(s, s * 0.6),
                QPointF(s, -s * 0.6),
                QPointF(-s, s * 0.6),
            ]
        )
        painter.drawPolygon(poly)
        painter.drawArc(QRectF(-s * 0.6, -s * 1.4, s * 1.2, s * 0.8), 0, 180 * 16)
        painter.drawLine(QPointF(0, -s * 0.1), QPointF(0, -s * 0.6))
    elif st == "Válvula_R":
        poly = QPolygonF(
            [
                QPointF(-s, -s * 0.6),
                QPointF(s, s * 0.6),
                QPointF(s, -s * 0.6),
                QPointF(-s, s * 0.6),
            ]
        )
        painter.drawPolygon(poly)
        painter.drawLine(QPointF(-s, s * 0.6), QPointF(s, -s * 0.6))
    elif st == "Chaminé":
        poly = QPolygonF(
            [
                QPointF(-s * 0.5, -s * 2),
                QPointF(s * 0.5, -s * 2),
                QPointF(s, s * 2),
                QPointF(-s, s * 2),
            ]
        )
        painter.drawPolygon(poly)
    elif st == "Refervedor":
        painter.drawRoundedRect(
            QRectF(-s * 1.2, -s * 0.8, s * 2.4, s * 1.6), s * 0.4, s * 0.4
        )
        painter.drawRect(QRectF(s * 0.2, -s * 0.4, s * 0.8, s * 0.8))
    elif st == "Extrusora":
        painter.drawRect(QRectF(-s * 2, -s * 0.4, s * 4, s * 0.8))
        painter.drawRect(QRectF(-s * 1.8, -s * 0.9, s * 0.6, s * 0.5))
        for i in range(8):
            painter.drawLine(
                QPointF(-s * 1.5 + i * s * 0.4, -s * 0.4),
                QPointF(-s * 1.3 + i * s * 0.4, s * 0.4),
            )
    elif st == "Reator_F":
        painter.drawRoundedRect(
            QRectF(-s * 0.8, -s * 2, s * 1.6, s * 4), s * 0.4, s * 0.4
        )
        for i in range(20):
            painter.drawPoint(
                QPointF(random.uniform(-s * 0.6, s * 0.6), random.uniform(s, s * 1.8))
            )
    elif st == "ESP":
        painter.drawRect(QRectF(-s * 1.5, -s, s * 3, s * 2))
        for i in range(4):
            painter.drawLine(
                QPointF(-s * 1.2 + i * s * 0.8, -s), QPointF(-s * 1.2 + i * s * 0.8, s)
            )
        poly = QPolygonF(
            [
                QPointF(-s * 1.5, s),
                QPointF(-s * 0.5, s * 1.5),
                QPointF(0, s),
                QPointF(0.5 * s, s * 1.5),
                QPointF(s * 1.5, s),
            ]
        )
        painter.drawPolyline(poly)
    elif st == "Filtro_R":
        painter.drawEllipse(QRectF(-s, -s, s * 2, s * 2))
        painter.drawRect(QRectF(-s * 1.2, s * 0.2, s * 2.4, s * 0.8))
    elif st == "Torre_R":
        painter.drawRoundedRect(
            QRectF(-s * 0.6, -s * 2.5, s * 1.2, s * 5), s * 0.5, s * 0.5
        )
        for i in range(3):
            y = -s * 1.5 + i * s * 1.5
            painter.drawRect(QRectF(-s * 0.5, y, s, s * 0.8))
            for j in range(3):
                painter.drawLine(
                    QPointF(-s * 0.5, y + j * 0.2 * s),
                    QPointF(s * 0.5, y + j * 0.2 * s + 0.1 * s),
                )
    elif st == "Válvula_S":
        poly = QPolygonF(
            [
                QPointF(-s, -s * 0.6),
                QPointF(s, s * 0.6),
                QPointF(s, -s * 0.6),
                QPointF(-s, s * 0.6),
            ]
        )
        painter.drawPolygon(poly)
        painter.drawRect(QRectF(-s * 0.3, -s * 1.3, s * 0.6, s * 0.7))
    elif st == "Tanque_F":
        painter.drawRect(QRectF(-s, -s, s * 2, s * 2))
        painter.drawLine(QPointF(-s, -s * 0.5), QPointF(s, -s * 0.5))
    elif st == "Peneira_V":
        painter.drawLine(QPointF(-s * 1.5, -s * 0.5), QPointF(s * 1.5, s * 0.5))
        for i in range(10):
            painter.drawPoint(QPointF(-s * 1.3 + i * s * 0.3, -s * 0.4 + i * s * 0.1))
    elif st == "Esfera":
        painter.drawEllipse(QRectF(-s, -s, s * 2, s * 2))
        painter.drawLine(QPointF(-s * 0.7, s), QPointF(-s * 0.7, s * 1.3))
        painter.drawLine(QPointF(s * 0.7, s), QPointF(s * 0.7, s * 1.3))
    elif st == "Correia":
        painter.drawRoundedRect(
            QRectF(-s * 2, -s * 0.2, s * 4, s * 0.4), s * 0.2, s * 0.2
        )
        painter.drawEllipse(QRectF(-s * 1.9, -s * 0.15, s * 0.3, s * 0.3))
        painter.drawEllipse(QRectF(s * 1.6, -s * 0.15, s * 0.3, s * 0.3))
    elif st == "Rosca":
        painter.drawRect(QRectF(-s * 2, -s * 0.3, s * 4, s * 0.6))
        for i in range(10):
            painter.drawLine(
                QPointF(-s * 2 + i * s * 0.4, -s * 0.3),
                QPointF(-s * 1.8 + i * s * 0.4, s * 0.3),
            )
    elif st == "Elevador":
        painter.drawRect(QRectF(-s * 0.4, -s * 2, s * 0.8, s * 4))
        for i in range(5):
            painter.drawRect(QRectF(-s * 0.3, -s * 1.8 + i * s * 0.8, s * 0.6, s * 0.3))
    elif st == "Heat_E":
        painter.drawRect(QRectF(-s, -s, s * 2, s * 2))
        painter.drawLine(QPointF(-s, -s), QPointF(s, s))
        painter.drawLine(QPointF(-s, s), QPointF(s, -s))
    elif st == "Flare":
        painter.drawLine(QPointF(-s * 0.3, s * 2), QPointF(0, -s * 2))
        painter.drawLine(QPointF(s * 0.3, s * 2), QPointF(0, -s * 2))
        painter.drawEllipse(QRectF(-s * 0.4, -s * 2.2, s * 0.8, s * 0.4))
    elif st == "Reator":
        painter.drawRoundedRect(
            QRectF(-s * 0.8, -s * 1.2, s * 1.6, s * 2.4), s * 0.4, s * 0.4
        )
        painter.drawRect(QRectF(-s * 0.25, -s * 1.6, s * 0.5, s * 0.4))
        painter.drawLine(QPointF(0, -s * 1.2), QPointF(0, s * 0.6))
        painter.drawLine(QPointF(-s * 0.4, s * 0.2), QPointF(s * 0.4, s * 0.6))
        painter.drawLine(QPointF(s * 0.4, s * 0.2), QPointF(-s * 0.4, s * 0.6))
    elif st == "Reator Tubular (PFR)":
        painter.drawRect(QRectF(-s * 1.6, -s * 0.5, s * 3.2, s))
        for i in [-1.0, -0.5, 0, 0.5, 1.0]:
            painter.drawEllipse(QRectF(s * i - s * 0.2, -s * 0.4, s * 0.4, s * 0.8))
    elif st == "Misturador":
        painter.drawEllipse(QRectF(-s, -s, s * 2, s * 2))
        painter.drawLine(QPointF(-s * 0.6, -s * 0.6), QPointF(s * 0.6, s * 0.6))
        painter.drawLine(QPointF(-s * 0.6, s * 0.6), QPointF(s * 0.6, -s * 0.6))
    elif st == "Moinho":
        painter.drawRect(QRectF(-s, -s * 0.5, s * 2, s))
        painter.drawEllipse(QRectF(-s * 0.8, -s * 0.4, s * 0.8, s * 0.8))
        painter.drawEllipse(QRectF(0, -s * 0.4, s * 0.8, s * 0.8))
    elif st == "Britador":
        poly = QPolygonF(
            [
                QPointF(-s * 0.8, -s),
                QPointF(s * 0.8, -s),
                QPointF(s * 0.4, s * 0.8),
                QPointF(-s * 0.4, s * 0.8),
            ]
        )
        painter.drawPolygon(poly)
        painter.drawLine(QPointF(0, -s), QPointF(0, s * 0.8))
    elif st == "Torre de Destilação":
        painter.drawRoundedRect(
            QRectF(-s * 0.6, -s * 2.5, s * 1.2, s * 5), s * 0.5, s * 0.5
        )
        for i in range(8):
            y = -s * 2 + i * s * 0.6
            painter.drawLine(QPointF(-s * 0.6, y), QPointF(s * 0.6, y))
    elif st == "Coluna de Absorção":
        painter.drawRoundedRect(
            QRectF(-s * 0.6, -s * 2.5, s * 1.2, s * 5), s * 0.5, s * 0.5
        )
        for i in [-1.5, -0.75, 0, 0.75, 1.5]:
            painter.drawLine(QPointF(-s * 0.6, s * i), QPointF(s * 0.6, s * i))
        painter.drawLine(QPointF(-s * 0.6, -s * 1.5), QPointF(s * 0.6, s * 1.5))
        painter.drawLine(QPointF(s * 0.6, -s * 1.5), QPointF(-s * 0.6, s * 1.5))
    elif st == "Secador":
        painter.drawRoundedRect(
            QRectF(-s * 1.2, -s * 0.6, s * 2.4, s * 1.2), s * 0.2, s * 0.2
        )
        painter.drawLine(QPointF(-s * 1.2, 0), QPointF(s * 1.2, 0))
    elif st == "Filtro Prensa":
        painter.drawRect(QRectF(-s * 1.2, -s * 0.8, s * 2.4, s * 1.6))
        for i in [-0.6, 0, 0.6]:
            painter.drawLine(QPointF(s * i, -s * 0.8), QPointF(s * i, s * 0.8))
    elif st == "Filtro Rotativo":
        painter.drawEllipse(QRectF(-s * 1.2, -s * 1.2, s * 2.4, s * 2.4))
        painter.drawRect(QRectF(-s * 1.5, s * 0.4, s * 3, s * 0.8))
    elif st == "Filtro":
        painter.drawEllipse(QRectF(-s, -s, s * 2, s * 2))
        painter.setPen(QPen(QColor(pen_color), 2, Qt.DashLine))
        painter.drawLine(QPointF(-s, 0), QPointF(s, 0))
        painter.setPen(default_pen)
    elif st == "Ciclone":
        painter.drawRect(QRectF(-s * 0.6, -s * 1.5, s * 1.2, s))
        poly = QPolygonF(
            [
                QPointF(-s * 0.6, -s * 0.5),
                QPointF(s * 0.6, -s * 0.5),
                QPointF(0, s * 1.5),
            ]
        )
        painter.drawPolygon(poly)
    elif st == "Hidrociclone":
        painter.drawRect(QRectF(-s * 0.4, -s * 1.5, s * 0.8, s))
        poly = QPolygonF(
            [
                QPointF(-s * 0.4, -s * 0.5),
                QPointF(s * 0.4, -s * 0.5),
                QPointF(0, s * 1.5),
            ]
        )
        painter.drawPolygon(poly)
    elif st == "Centrífuga":
        painter.drawEllipse(QRectF(-s * 1.2, -s * 0.8, s * 2.4, s * 1.6))
        painter.drawLine(QPointF(-s * 1.2, 0), QPointF(s * 1.2, 0))
    elif st == "Clarificador":
        painter.drawRect(QRectF(-s * 1.5, -s * 0.5, s * 3, s))
        poly = QPolygonF(
            [
                QPointF(-s * 1.5, s * 0.5),
                QPointF(s * 1.5, s * 0.5),
                QPointF(s * 0.5, s * 1.5),
                QPointF(-s * 0.5, s * 1.5),
            ]
        )
        painter.drawPolygon(poly)
    elif st == "Peneira Vibratória" or st == "Peneira":
        painter.drawRect(QRectF(-s * 1.2, -s * 0.8, s * 2.4, s * 1.6))
        painter.setPen(QPen(QColor(pen_color), 2, Qt.DashLine))
        painter.drawLine(QPointF(-s * 1.2, -s * 0.4), QPointF(s * 1.2, s * 0.4))
        painter.setPen(default_pen)
    elif st == "Válvula":
        poly = QPolygonF(
            [
                QPointF(-s, -s * 0.5),
                QPointF(-s, s * 0.5),
                QPointF(s, -s * 0.5),
                QPointF(s, s * 0.5),
            ]
        )
        painter.drawPolygon(poly)
    elif st == "Válvula Borboleta":
        poly = QPolygonF(
            [
                QPointF(-s, -s * 0.5),
                QPointF(-s, s * 0.5),
                QPointF(s, -s * 0.5),
                QPointF(s, s * 0.5),
            ]
        )
        painter.drawPolygon(poly)
        painter.drawEllipse(QRectF(-s * 0.2, -s * 0.2, s * 0.4, s * 0.4))
    elif st == "Válvula de Controle":
        poly = QPolygonF(
            [
                QPointF(-s, -s * 0.5),
                QPointF(-s, s * 0.5),
                QPointF(s, -s * 0.5),
                QPointF(s, s * 0.5),
            ]
        )
        painter.drawPolygon(poly)
        painter.drawLine(QPointF(0, 0), QPointF(0, -s * 1.2))
        painter.drawEllipse(QRectF(-s * 0.6, -s * 1.8, s * 1.2, s * 0.6))
    elif st == "Válvula de Retenção":
        poly = QPolygonF(
            [
                QPointF(-s, -s * 0.5),
                QPointF(-s, s * 0.5),
                QPointF(s, -s * 0.5),
                QPointF(s, s * 0.5),
            ]
        )
        painter.drawPolygon(poly)
        painter.drawLine(QPointF(0, -s * 0.5), QPointF(0, s * 0.5))
    elif st == "PSV":
        poly = QPolygonF(
            [
                QPointF(-s, -s * 0.5),
                QPointF(-s, s * 0.5),
                QPointF(s, -s * 0.5),
                QPointF(s, s * 0.5),
            ]
        )
        painter.drawPolygon(poly)
        painter.drawLine(QPointF(0, 0), QPointF(0, -s))
        painter.drawLine(QPointF(-s * 0.4, -s * 0.6), QPointF(s * 0.4, -s * 0.6))
    elif st == "Chaminé":
        poly = QPolygonF(
            [
                QPointF(-s * 0.4, -s * 2),
                QPointF(s * 0.4, -s * 2),
                QPointF(s * 0.8, s * 2),
                QPointF(-s * 0.8, s * 2),
            ]
        )
        painter.drawPolygon(poly)
    elif st == "Forno":
        painter.setBrush(QBrush(QColor(_bg_node)))
        painter.drawRoundedRect(QRectF(-s * 1.2, -s * 0.8, s * 2.4, s * 1.6), 10, 10)
        path = QPainterPath()
        path.moveTo(-s * 0.8, s * 0.2)
        for i in range(5):
            path.lineTo(-s * 0.6 + i * s * 0.3, -s * 0.2 if i % 2 == 0 else s * 0.2)
        painter.setBrush(Qt.NoBrush)
        painter.drawPath(path)
    elif st == "Lavadora":
        painter.setBrush(QBrush(QColor(_bg_node)))
        painter.drawRoundedRect(QRectF(-s * 0.7, -s * 1.3, s * 1.4, s * 2.6), 15, 15)
        for i in range(4):
            painter.drawLine(
                QPointF(-s * 0.5, -s + i * s * 0.6), QPointF(s * 0.5, -s + i * s * 0.6)
            )
    elif st == "Box":
        painter.setBrush(QBrush(QColor(_bg_node)))
        painter.setPen(QPen(QColor(theme["accent"]), 1.5))
        painter.drawRect(QRectF(-s * 0.8, -s * 0.8, s * 1.6, s * 1.6))
        painter.drawLine(QPointF(-s * 0.8, -s * 0.8), QPointF(s * 0.8, s * 0.8))
        painter.drawLine(QPointF(-s * 0.8, s * 0.8), QPointF(s * 0.8, -s * 0.8))
    else:
        painter.setBrush(QBrush(QColor(_bg_node)))
        painter.setPen(QPen(QColor(theme["accent"]), 2, Qt.DashLine))
        painter.drawRect(QRectF(-s, -s, s * 2, s * 2))
        painter.drawLine(QPointF(-s, -s), QPointF(s, s))
        painter.drawLine(QPointF(-s, s), QPointF(s, -s))


class ConnectorPort(QGraphicsEllipseItem):
    def __init__(self, node, port_id):
        super().__init__(-5, -5, 10, 10, node)
        self.node = node
        self.port_id = port_id
        self.setAcceptHoverEvents(True)
        self.setBrush(QBrush(QColor(0, 150, 255)))
        self.setPen(QPen(QColor(255, 255, 255), 1))
        self.setCursor(QCursor(Qt.CrossCursor))
        self.setZValue(10)
        self.setVisible(False)

    def hoverEnterEvent(self, event):
        self.setBrush(QBrush(QColor(0, 255, 255)))
        super().hoverEnterEvent(event)

    def hoverLeaveEvent(self, event):
        self.setBrush(QBrush(QColor(0, 150, 255)))
        super().hoverLeaveEvent(event)

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            view = self.scene().views()[0]
            if hasattr(view, "start_connection"):
                view.start_connection(self)
                event.accept()
                return
        super().mousePressEvent(event)


class SourceSinkHandle(QGraphicsItem):
    """Ponta de seta independente (Source/Sink) que não depende de equipamentos."""

    def __init__(self, h_type="Saída", flow_name=""):
        super().__init__()
        self.h_type = h_type
        self.flow_name = flow_name
        self.setFlags(
            QGraphicsItem.ItemIsMovable
            | QGraphicsItem.ItemIsSelectable
            | QGraphicsItem.ItemSendsGeometryChanges
        )
        self.edges = []
        # Adiciona uma porta real para permitir reconectar arrastando
        self.ports = {"tip": ConnectorPort(self, "tip")}
        self.ports["tip"].setPos(0, 0)
        self.ports["tip"].setBrush(
            QBrush(QColor(0, 255, 255))
        )  # Cor ciano para terminais
        self.ports["tip"].setVisible(False)

    def add_edge(self, edge):
        if edge not in self.edges:
            self.edges.append(edge)

    def boundingRect(self):
        return QRectF(-20, -20, 40, 40)

    def paint(self, painter, option, widget=None):
        # Tornar o terminal invisível — Queremos apenas a SEÇÃO da seta do Edge
        pass

    def contextMenuEvent(self, event):
        menu = QMenu()
        t = T()
        menu.setStyleSheet(
            f"QMenu{{background:{t['bg_card']};color:{t['text']};border:1px solid {t['accent']};}}"
        )
        name_act = menu.addAction("✏️ Renomear Terminal")
        del_act = menu.addAction("🗑 Excluir Terminal")
        act = menu.exec_(event.screenPos())
        if act == name_act:
            text, ok = QInputDialog.getText(
                None, "Terminal", "Nome:", QLineEdit.Normal, self.flow_name
            )
            if ok:
                self.flow_name = text
                self.update()
        elif act == del_act:
            for e in list(self.edges):
                if e.scene():
                    self.scene().removeItem(e)
            if self.scene():
                self.scene().removeItem(self)

    def itemChange(self, change, value):
        if change == QGraphicsItem.ItemPositionHasChanged:
            for e in self.edges:
                e.adjust()
        return super().itemChange(change, value)


class JunctionNode(QGraphicsEllipseItem):
    """Nó de interconexão minimalista para T-junctions entre tubulações."""

    def __init__(self):
        super().__init__(-6, -6, 12, 12)
        self.setFlags(
            QGraphicsItem.ItemIsMovable
            | QGraphicsItem.ItemIsSelectable
            | QGraphicsItem.ItemSendsGeometryChanges
        )
        self.edges = []
        self.ports = {"tip": ConnectorPort(self, "tip")}
        self.ports["tip"].setPos(0, 0)
        self.ports["tip"].setBrush(QBrush(QColor(0, 255, 255)))  # Ciano Neon
        self.ports["tip"].setVisible(False)
        self._bg = T()["accent"]
        self.custom_name = "Junção"
        # --- Armazenamento de Massa para Balanço Estrito ---
        self.stored_components = {}  # {componente: massa_acumulada}
        self.total_input = {}  # {componente: total_entrada}
        self.total_output = {}  # {componente: total_saida}
        self.mass_balance_error = 0.0  # Erro de fechamento do nó

    def add_edge(self, edge):
        if edge not in self.edges:
            self.edges.append(edge)

    def paint(self, painter, option, widget=None):
        t = T()
        painter.setPen(QPen(QColor(t["accent"]), 2))
        painter.setBrush(QBrush(QColor(t["bg_app"])))
        painter.drawEllipse(self.rect())

        # Ponto central se selecionado ou hover
        if self.isSelected():
            painter.setBrush(QBrush(QColor(t["accent_bright"])))
            painter.drawEllipse(-3, -3, 6, 6)

    def contextMenuEvent(self, event):
        menu = QMenu()
        t = T()
        menu.setStyleSheet(
            f"QMenu{{background:{t['bg_card']};color:{t['text']};border:1px solid {t['accent']};}}"
        )
        del_act = menu.addAction("🗑 Excluir Junção")
        action = menu.exec_(event.screenPos())
        if action == del_act:
            for e in list(self.edges):
                if e.scene():
                    self.scene().removeItem(e)
            if self.scene():
                self.scene().removeItem(self)

    def itemChange(self, change, value):
        if change == QGraphicsItem.ItemPositionHasChanged:
            for e in self.edges:
                e.adjust()
        return super().itemChange(change, value)


class Edge(QGraphicsPathItem):
    def __init__(self, source_node, dest_node, source_port="right", dest_port="left"):
        super().__init__()
        self.source_node = source_node
        self.dest_node = dest_node
        self.source_port = source_port
        self.dest_port = dest_port
        self.source_node.add_edge(self)
        self.dest_node.add_edge(self)
        self.setFlag(QGraphicsItem.ItemIsSelectable)
        try:
            lc = T()["line"]
        except:
            lc = "#94a3b8"
        pen = QPen(QColor(lc), 1.5, Qt.SolidLine, Qt.RoundCap, Qt.RoundJoin)
        self.setPen(pen)
        self.setZValue(-1)
        self.is_utility = False
        self.pipe_name = ""

        # --- Dados do Balanço de Massa Avançado ---
        self.flow_data = {}  # ex: {"Água": 100.0}

        # Container para o label com fundo para visibilidade
        self.label_bg = QGraphicsRectItem(self)
        self.label_bg.setBrush(QBrush(QColor(T()["bg_app"])))
        self.label_bg.setPen(QPen(Qt.NoPen))
        self.label_bg.setVisible(False)

        self.label_item = QGraphicsSimpleTextItem(self.label_bg)
        self.label_item.setBrush(QBrush(QColor(T()["text"])))
        self.label_item.setFont(QFont("Segoe UI", 8, QFont.Bold))
        self.label_item.setVisible(True)

        self.adjust()

    def get_port_normal(self, port_id):
        side = port_id.split("_")[0]
        if side == "top":
            return (0, -1)
        if side == "bottom":
            return (0, 1)
        if side == "left":
            return (-1, 0)
        if side == "right":
            return (1, 0)
        return (0, 0)

    def get_port_position(self, node, port_id):
        if isinstance(node, SourceSinkHandle):
            return node.scenePos()
        if hasattr(node, "ports") and port_id in node.ports:
            return node.mapToScene(node.ports[port_id].pos())
        return node.sceneBoundingRect().center()

    def adjust(self):
        if not self.source_node or not self.dest_node:
            return
        p1 = self.get_port_position(self.source_node, self.source_port)
        p2 = self.get_port_position(self.dest_node, self.dest_port)
        path = QPainterPath()
        path.moveTo(p1)
        d1 = self.get_port_normal(self.source_port)
        d2 = self.get_port_normal(self.dest_port)
        p1_out = QPointF(p1.x() + d1[0] * 25, p1.y() + d1[1] * 25)
        p2_in = QPointF(p2.x() + d2[0] * 25, p2.y() + d2[1] * 25)
        path.lineTo(p1_out)
        if d1[1] != 0:
            path.lineTo(p1_out.x(), p1_out.y())
            path.lineTo(p2_in.x(), p1_out.y())
            path.lineTo(p2_in.x(), p2_in.y())
        else:
            path.lineTo(p1_out.x(), p2_in.y())
            path.lineTo(p2_in.x(), p2_in.y())
        path.lineTo(p2)
        self.setPath(path)

        # Formata o texto para exibir dados do balanço junto com o nome
        total_flow = sum(self.flow_data.values())
        display_text = self.pipe_name
        if total_flow > 0:
            display_text += f"\n{total_flow:.1f} kg/h"

        if display_text:
            # Encontra o ponto médio do caminho para o label
            mid_p = path.pointAtPercent(0.5)
            self.label_item.setText(display_text)
            self.label_item.setBrush(QBrush(QColor(T()["text"])))
            self.label_bg.setBrush(QBrush(QColor(T()["bg_app"])))
            br = self.label_item.boundingRect()
            self.label_bg.setRect(0, 0, br.width() + 8, br.height() + 4)
            self.label_item.setPos(4, 2)
            self.label_bg.setPos(
                mid_p.x() - br.width() / 2 - 4, mid_p.y() - br.height() - 10
            )
            self.label_bg.setVisible(True)
            self.label_bg.setZValue(1)
        else:
            self.label_bg.setVisible(False)
        dx, dy = p2.x() - p2_in.x(), p2.y() - p2_in.y()
        self._arrow_dir = math.atan2(dy, dx)
        self._dest_pos = p2

    def contextMenuEvent(self, event):
        menu = QMenu()
        t = T()
        menu.setStyleSheet(
            f"QMenu{{background:{t['bg_card']};color:{t['text']};border:1px solid {t['accent']};}}"
        )
        data_act = menu.addAction("📊 Editar Dados da Corrente")
        name_action = menu.addAction("✏️ Renomear Fluxo")
        del_act = menu.addAction("🗑 Excluir Tubulação")
        action = menu.exec_(event.screenPos())
        if action == del_act:
            for n in [self.source_node, self.dest_node]:
                if isinstance(n, SourceSinkHandle):
                    if n.scene():
                        self.scene().removeItem(n)
                elif n and self in n.edges:
                    n.edges.remove(self)
            if self.scene():
                self.scene().removeItem(self)
        elif action == name_action:
            text, ok = QInputDialog.getText(
                None, "Piping", "Nome:", QLineEdit.Normal, self.pipe_name
            )
            if ok:
                self.pipe_name = text
                self.label_item.setText(text)
                self.adjust()
        elif action == data_act:
            parent_widget = self.scene().views()[0].parentWidget()
            global_comps = getattr(
                parent_widget, "global_components", INDUSTRIAL_COMPONENTS
            )
            dialog = StreamEditorDialog(self, global_components=global_comps)
            dialog.exec_()

    def paint(self, painter, option, widget=None):
        try:
            t = T()
            lc = t.get("line", C_LINE)
        except:
            t = {"accent_bright": "#00E5FF"}
            lc = C_LINE

        style = Qt.SolidLine
        painter.setRenderHint(QPainter.Antialiasing)

        if self.isSelected():
            painter.setPen(
                QPen(
                    QColor(t.get("accent_bright", "#00E5FF")),
                    3,
                    style,
                    Qt.RoundCap,
                    Qt.RoundJoin,
                )
            )
        else:
            painter.setPen(QPen(QColor(lc), 2.5, style, Qt.RoundCap, Qt.RoundJoin))

        painter.drawPath(self.path())

        if hasattr(self, "_dest_pos"):
            color = (
                QColor(lc)
                if not self.isSelected()
                else QColor(t.get("accent_bright", "#34d399"))
            )
            painter.setPen(QPen(Qt.NoPen))
            painter.setBrush(QBrush(color))

            p1, angle = self._dest_pos, self._arrow_dir
            size = 8

            # Seta Sólida Industrial (Triângulo)
            p2 = QPointF(
                p1.x() - size * math.cos(angle - 0.35),
                p1.y() - size * math.sin(angle - 0.35),
            )
            p3 = QPointF(
                p1.x() - size * math.cos(angle + 0.35),
                p1.y() - size * math.sin(angle + 0.35),
            )

            arrow_head = QPolygonF([p1, p2, p3])
            painter.drawPolygon(arrow_head)


class ProcessNode(QGraphicsItem):
    def __init__(self, symbol_type):
        super().__init__()
        self.symbol_type = symbol_type
        self.edges = []
        self.size = 50
        self.setFlags(
            QGraphicsItem.ItemIsMovable
            | QGraphicsItem.ItemIsSelectable
            | QGraphicsItem.ItemSendsGeometryChanges
        )
        self.setAcceptHoverEvents(True)
        self.setCursor(QCursor(Qt.SizeAllCursor))
        self.ports = {
            f"{s}_{i}": ConnectorPort(self, f"{s}_{i}")
            for s in ["top", "bottom", "left", "right"]
            for i in range(1, 4)
        }
        self.update_ports()
        self.custom_name = ""
        self._hover_bg = QGraphicsRectItem(self)
        self._hover_bg.setBrush(QBrush(QColor(20, 20, 20, 230)))
        self._hover_bg.setPen(QPen(QColor(T()["accent"]), 1))
        self._hover_text = QGraphicsSimpleTextItem(self.symbol_type, self)
        self._hover_text.setBrush(QBrush(QColor("white")))
        self._hover_text.setFont(QFont("Segoe UI", 9, QFont.Bold))
        tw, th = (
            self._hover_text.boundingRect().width(),
            self._hover_text.boundingRect().height(),
        )
        self._hover_bg.setRect(-5, -5, tw + 10, th + 10)
        lx, ly = -tw / 2, -self.size / 2 - th - 15
        self._hover_bg.setPos(lx, ly)
        self._hover_text.setPos(lx, ly)
        self._hover_bg.setVisible(False)
        self._hover_text.setVisible(False)
        self._hover_bg.setAcceptHoverEvents(False)
        self._hover_text.setAcceptHoverEvents(False)

        # Rótulo Persistente (Nome do Equipamento) — Com quebra de linha (Word Wrap)
        self._name_item = QGraphicsTextItem(self)
        self._name_item.setDefaultTextColor(QColor(T()["text"]))
        self._name_item.setFont(QFont("Segoe UI", 9, QFont.Bold))
        self._name_item.setTextWidth(120)  # Largura máxima para forçar quebra de linha
        opt = self._name_item.document().defaultTextOption()
        opt.setAlignment(Qt.AlignCenter)
        self._name_item.document().setDefaultTextOption(opt)

        self.update_name_pos()

        # --- Configuração de Balanço de Massa (Split) ---
        # { component: { port_id: {mode: "perc"/"fixed", val: float} } }
        self.split_config = {}
        # --- Armazenamento de Massa para Balanço Estrito ---
        self.stored_components = {}  # {componente: massa_acumulada_no_equipamento}
        self.total_input = {}  # {componente: total_entrada_acumulada}
        self.total_output = {}  # {componente: total_saida_acumulada}
        self.mass_balance_error = 0.0  # Erro de fechamento do equipamento
        self.has_steady_state = False  # Indica se atingiu estado estacionário

    def update_name_pos(self):
        # r é a área central do equipamento, ignorando os paddings do boundingRect
        # Vamos posicionar o texto de forma fixa em relação ao centro (0,0) do nó
        tw = self._name_item.boundingRect().width()
        th = self._name_item.boundingRect().height()
        # Posiciona abaixo do ícone (o ícone costuma ficar em torno de self.size)
        self._name_item.setPos(-tw / 2, self.size / 2 + 5)

    def hoverEnterEvent(self, event):
        for port in self.ports.values():
            port.setVisible(True)
        self.prepareGeometryChange()
        self._hover_bg.setVisible(True)
        self._hover_text.setVisible(True)
        super().hoverEnterEvent(event)

    def hoverLeaveEvent(self, event):
        for port in self.ports.values():
            port.setVisible(False)
        self._hover_bg.setVisible(False)
        self._hover_text.setVisible(False)
        super().hoverLeaveEvent(event)

    def update_ports(self):
        s = self.size / 2
        st = self._get_alias()
        m = 0
        if st in ["Torre de Destilação", "Coluna de Absorção", "Flare", "Chaminé"]:
            r = QRectF(-s * 0.8 - m, -s * 2.7 - m, s * 1.6 + m * 2, s * 5.4 + m * 2)
        elif st in [
            "Separador Bifásico",
            "Separador Trifásico",
            "Caldeira",
            "Vaso Horizontal",
            "Clarificador",
        ]:
            r = QRectF(-s * 1.8 - m, -s - m, self.size * 1.8 + m * 2, self.size + m * 2)
        elif st in [
            "Secador",
            "Peneira",
            "Peneira Vibratória",
            "Filtro Prensa",
            "Trocador de Placas",
            "Filtro Rotativo",
        ]:
            r = QRectF(
                -s * 1.5 - m,
                -s * 1.2 - m,
                self.size * 1.5 + m * 2,
                self.size * 1.2 + m * 2,
            )
        elif st in [
            "Reator",
            "Reator Tubular (PFR)",
            "Ciclone",
            "Hidrociclone",
            "Torre de Resfriamento",
            "Silo",
        ]:
            r = QRectF(
                -s * 1.8 - m,
                -s * 1.8 - m,
                self.size * 1.8 + m * 2,
                self.size * 1.8 + m * 2,
            )
        elif st in ["Evaporador", "Aquecedor Elétrico"]:
            r = QRectF(-s - m, -s * 1.6 - m, self.size + m * 2, self.size * 1.6 + m * 2)
        else:
            r = QRectF(-s - m, -s - m, self.size + m * 2, self.size + m * 2)
        if hasattr(self, "ports"):
            for i in range(1, 4):
                f = i * 0.25
                self.ports[f"top_{i}"].setPos(r.left() + r.width() * f, r.top())
                self.ports[f"bottom_{i}"].setPos(r.left() + r.width() * f, r.bottom())
                self.ports[f"left_{i}"].setPos(r.left(), r.top() + r.height() * f)
                self.ports[f"right_{i}"].setPos(r.right(), r.top() + r.height() * f)

    def _get_alias(self):
        return EQUIPMENT_ALIASES.get(self.symbol_type, self.symbol_type)

    def add_edge(self, edge):
        if edge not in self.edges:
            self.edges.append(edge)

    def set_name(self, name):
        self.custom_name = name
        self._hover_text.setText(
            f"{self.symbol_type}: {name}" if name else self.symbol_type
        )
        self._name_item.setPlainText(name if name else self.symbol_type)
        self.update_name_pos()
        # Re-ajustar fundo do hover
        tw, th = (
            self._hover_text.boundingRect().width(),
            self._hover_text.boundingRect().height(),
        )
        self._hover_bg.setRect(-5, -5, tw + 10, th + 10)
        self.update()

    def set_size(self, new_size):
        self.prepareGeometryChange()
        self.size = max(20, min(new_size, 300))
        self.update_ports()
        for e in self.edges:
            e.adjust()
        self.update()

    def contextMenuEvent(self, event):
        menu = QMenu()
        t = T()
        menu.setStyleSheet(
            f"QMenu{{background:{t['bg_card']};color:{t['text']};border:1px solid {t['accent']};}}"
        )
        perf_act = menu.addAction("⚙️ Configurar Desempenho")
        del_act = menu.addAction("🗑 Excluir Equipamento")
        act = menu.exec_(event.screenPos())
        if act == del_act:
            for e in list(self.edges):
                other = e.source_node if e.dest_node == self else e.dest_node
                if isinstance(other, SourceSinkHandle):
                    if other.scene():
                        self.scene().removeItem(other)
                if e.scene():
                    self.scene().removeItem(e)
            if self.scene():
                self.scene().removeItem(self)
        elif act == perf_act:
            EquipmentEditorDialog(self).exec_()

    def mouseDoubleClickEvent(self, event):
        EquipmentEditorDialog(self).exec_()
        super().mouseDoubleClickEvent(event)

    def itemChange(self, change, value):
        if change == QGraphicsItem.ItemPositionHasChanged:
            for e in self.edges:
                e.adjust()
        return super().itemChange(change, value)

    def boundingRect(self):
        m, s, st = 5, self.size / 2, self._get_alias()
        if st in ["Torre de Destilação", "Coluna de Absorção", "Flare", "Chaminé"]:
            base = QRectF(-s * 0.8 - m, -s * 2.7 - m, s * 1.6 + m * 2, s * 5.4 + m * 2)
        elif st in [
            "Separador Bifásico",
            "Separador Trifásico",
            "Caldeira",
            "Vaso Horizontal",
            "Clarificador",
        ]:
            base = QRectF(
                -s * 1.8 - m, -s - m, self.size * 1.8 + m * 2, self.size + m * 2
            )
        else:
            base = QRectF(-s - m, -s - m, self.size + m * 2, self.size + m * 2)
        return base.adjusted(-s * 2, -60, s * 2, s * 2)

    def paint(self, painter, option, widget=None):
        t = T()
        painter.setRenderHint(QPainter.Antialiasing)
        if self.isSelected():
            painter.setPen(QPen(QColor(t["accent_bright"]), 2, Qt.DotLine))
            painter.drawRect(self.boundingRect().adjusted(0, 40, 0, 0))
        # Sincroniza cor do texto com o tema no redesenho
        self._name_item.setDefaultTextColor(QColor(t["text"]))
        draw_equipment(painter, self.symbol_type, self.size, False, t)


class SymbolListWidget(QListWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setDragDropMode(QListWidget.DragOnly)
        self.setViewMode(QListWidget.ListMode)
        self.setIconSize(QSize(50, 50))
        self.setSpacing(5)
        self.setWordWrap(True)

    def startDrag(self, actions):
        item = self.currentItem()
        if not item:
            return
        data = QByteArray()
        QDataStream(data, QIODevice.WriteOnly).writeQString(item.text())
        mime = QMimeData()
        mime.setData("application/x-pfd-item", data)
        drag = QDrag(self)
        drag.setMimeData(mime)
        pix = item.icon().pixmap(40, 40)
        drag.setPixmap(pix)
        drag.setHotSpot(QPoint(20, 20))
        drag.exec_(Qt.CopyAction)


class SymbolPalette(QToolBox):
    def __init__(self):
        super().__init__()
        self.categories = {
            "Geral": [
                "Filtro",
                "Peneira",
                "Separador Bifásico",
                "Clarificador",
                "Secador Rotativo",
            ],
            "Armazenamento": [
                "Tanque Aberto",
                "Tanque Fechado",
                "Vaso Vertical",
                "Vaso Horizontal",
                "Esfera de Gás",
                "Silo",
                "Tanque de Teto Flutuante",
                "LPG Esfera",
            ],
            "Movimentação": [
                "Bomba Centrífuga",
                "Bomba Volumétrica",
                "Compressor",
                "Soprador",
                "Exaustor",
                "Turbina",
                "Ejetor",
                "Correia Transportadora",
                "Rosca Transportadora",
                "Elevador de Canecas",
            ],
            "Troca Térmica": [
                "Trocador Casco-Tubo",
                "Trocador de Placas",
                "Forno",
                "Caldeira",
                "Torre de Resfriamento",
                "Evaporador",
                "Aquecedor Elétrico",
                "Refervedor Kettle",
                "Condensador de Casco",
            ],
            "Transformação": [
                "Reator CSTR",
                "Reator Tubular (PFR)",
                "Leito Fluidizado",
                "Lavagem",
                "Tanque Misturador",
                "Moinho",
                "Britador",
                "Extrusora",
            ],
            "Separação": [
                "Torre de Destilação",
                "Coluna de Absorção",
                "Coluna de Recheio",
                "Decantador",
                "Hidrociclone",
                "Precipitador Eletrostático",
                "Filtro Rotativo",
                "Filtro de Mangas",
                "Filtro Prensa",
                "Cristalizador",
            ],
            "Instrumentação": [
                "Transmissor de Pressão",
                "Transmissor de Temperatura",
                "Transmissor de Vazão",
                "Transmissor de Nível",
                "Analisador",
            ],
            "Controles": [
                "Válvula Globo",
                "Válvula Borboleta",
                "Válvula de Esfera",
                "Válvula de Controle",
                "Válvula de Retenção",
                "PSV (Alívio)",
                "Válvula Solenoide",
                "Flare",
                "Chaminé",
                "Painel de Controle",
            ],
        }
        for cat, symbols in self.categories.items():
            lw = SymbolListWidget()
            self._apply_list_style(lw)
            for sym in symbols:
                item = QListWidgetItem(self._create_icon(sym), sym)
                item.setToolTip(f"Equipamento: {sym}")
                lw.addItem(item)
            self.addItem(lw, cat)
        self._apply_palette_style()

    def _create_icon(self, sym):
        pix = QPixmap(50, 50)
        pix.fill(Qt.transparent)
        p = QPainter(pix)
        p.setRenderHint(QPainter.Antialiasing)
        p.translate(25, 25)
        draw_equipment(p, sym, 28, True, T())
        p.end()
        return QIcon(pix)

    def refresh_theme(self):
        self._apply_palette_style()
        for i in range(self.count()):
            lw = self.widget(i)
            if isinstance(lw, SymbolListWidget):
                self._apply_list_style(lw)
                # Atualiza ícones de cada item
                for row in range(lw.count()):
                    item = lw.item(row)
                    item.setIcon(self._create_icon(item.text()))

    def _apply_palette_style(self):
        t = T()
        self.setStyleSheet(f"""
            QToolBox {{ 
                background-color: {t["bg_card"]}; 
                border: none;
            }}
            QToolBox::tab {{
                background-color: {t["bg_card"]};
                color: {t["text"]};
                font-weight: bold;
                font-size: 13px;
                font-family: 'Segoe UI';
                border-bottom: 1px solid {t["accent_dim"]};
                padding: 8px 12px;
            }}
            QToolBox::tab:selected {{
                color: {t["accent"]};
                border-bottom: 2px solid {t["accent"]};
            }}
        """)

    def _apply_list_style(self, lw):
        t = T()
        lw.setStyleSheet(f"""
            QListWidget {{ background: {t["bg_app"]}; border: none; padding: 5px; }}
            QListWidget::item {{ color: {t["text"]}; font-size: 11px; font-weight: bold; border-bottom: 1px solid {t["accent_dim"]}; padding: 10px; }}
            QListWidget::item:hover {{ background: {t["bg_card"]}; border-radius: 8px; color: {t["accent_bright"]}; }}
        """)


class FlowsheetCanvas(QGraphicsView):
    def __init__(self, scene):
        super().__init__(scene)
        self.setRenderHint(QPainter.Antialiasing)
        self.setAcceptDrops(True)
        self.mode = "Move"
        self.temp_line = None
        self.zoom_level = 1.0

    def dragEnterEvent(self, e):
        if e.mimeData().hasFormat("application/x-pfd-item"):
            e.acceptProposedAction()

    def dragMoveEvent(self, e):
        if e.mimeData().hasFormat("application/x-pfd-item"):
            e.acceptProposedAction()

    def dropEvent(self, e):
        sym = QDataStream(
            e.mimeData().data("application/x-pfd-item"), QIODevice.ReadOnly
        ).readQString()
        node = ProcessNode(sym)
        node.setPos(self.mapToScene(e.pos()))
        self.scene().addItem(node)
        e.acceptProposedAction()

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            items = self.items(event.pos())
            # Prioridade para ConnectorPort
            port = next((i for i in items if isinstance(i, ConnectorPort)), None)
            if port:
                self.mode, self.start_port = "PortConnect", port
                path = QPainterPath()
                path.moveTo(port.scenePos())
                self.temp_line = self.scene().addPath(
                    path, QPen(QColor(T()["accent"]), 2, Qt.DashLine)
                )
                # Mostra todas as portas para conexão
                for i in self.scene().items():
                    if hasattr(i, "ports"):
                        for p in i.ports.values():
                            p.setVisible(True)
                return

            # Se clicou em um Terminal que já tem porta escondida, mostra ela
            handle = next((i for i in items if isinstance(i, SourceSinkHandle)), None)
            if handle:
                handle.ports["tip"].setVisible(True)

        super().mousePressEvent(event)

    def mouseMoveEvent(self, event):
        if self.mode == "PortConnect" and self.temp_line:
            p1, p2 = self.start_port.scenePos(), self.mapToScene(event.pos())
            path = QPainterPath()
            path.moveTo(p1)

            # Lógica Ortogonal (L-shape) para o Preview
            # Similar ao Edge.adjust()
            d1 = (0, 0)
            if hasattr(self.start_port.node, "ports"):  # Se for de um ProcessNode
                for pid, po in self.start_port.node.ports.items():
                    if po == self.start_port:
                        side = pid.split("_")[0]
                        if side == "top":
                            d1 = (0, -1)
                        elif side == "bottom":
                            d1 = (0, 1)
                        elif side == "left":
                            d1 = (-1, 0)
                        elif side == "right":
                            d1 = (1, 0)
                        break

            p1_out = QPointF(p1.x() + d1[0] * 25, p1.y() + d1[1] * 25)
            path.lineTo(p1_out)
            if d1[1] != 0:  # Norte/Sul
                path.lineTo(p2.x(), p1_out.y())
            else:  # Leste/Oeste ou Terminal
                path.lineTo(p1_out.x(), p2.y())
            path.lineTo(p2)

            self.temp_line.setPath(path)
            return
        super().mouseMoveEvent(event)

    def mouseReleaseEvent(self, event):
        if self.mode != "PortConnect" or not self.temp_line:
            super().mouseReleaseEvent(event)
            return

        target = None
        items = self.items(event.pos())
        target = next(
            (
                i
                for i in items
                if isinstance(i, ConnectorPort) and i.node != self.start_port.node
            ),
            None,
        )

        target_edge = None
        if not target:
            target_edge = next(
                (
                    i
                    for i in items
                    if isinstance(i, Edge)
                    and i.source_node != self.start_port.node
                    and i.dest_node != self.start_port.node
                ),
                None,
            )

        if target:
            edge = Edge(
                self.start_port.node,
                target.node,
                self.start_port.port_id,
                target.port_id,
            )
            edge.flow_data = self._get_outlet_flow_data(self.start_port.node)
            self.scene().addItem(edge)
            edge.adjust()
            self.scene().removeItem(self.temp_line)
            self.temp_line = None
            self.mode = "Move"
            for i in self.scene().items():
                if hasattr(i, "ports"):
                    for p in i.ports.values():
                        p.setVisible(False)
            parent_widget = self.scene().views()[0].parentWidget()
            global_comps = getattr(
                parent_widget, "global_components", INDUSTRIAL_COMPONENTS
            )
            StreamEditorDialog(edge, global_components=global_comps).exec_()
        elif target_edge:
            pos = self.mapToScene(event.pos())
            junction = JunctionNode()
            junction.setPos(pos)
            self.scene().addItem(junction)

            old_flow = target_edge.flow_data.copy()
            old_name = target_edge.pipe_name
            src_n, dst_n = target_edge.source_node, target_edge.dest_node
            src_p, dst_p = target_edge.source_port, target_edge.dest_port

            self.scene().removeItem(target_edge)
            if target_edge in src_n.edges:
                src_n.edges.remove(target_edge)
            if target_edge in dst_n.edges:
                dst_n.edges.remove(target_edge)

            e1 = Edge(src_n, junction, src_p, "tip")
            e1.flow_data = old_flow
            e1.pipe_name = old_name
            self.scene().addItem(e1)
            e1.adjust()

            e2 = Edge(junction, dst_n, "tip", dst_p)
            e2.flow_data = old_flow
            e2.pipe_name = f"{old_name} (cont.)" if old_name else ""
            self.scene().addItem(e2)
            e2.adjust()

            new_leg = Edge(
                self.start_port.node, junction, self.start_port.port_id, "tip"
            )
            new_leg.flow_data = self._get_outlet_flow_data(self.start_port.node)
            self.scene().addItem(new_leg)
            new_leg.adjust()

            self.scene().removeItem(self.temp_line)
            self.temp_line = None
            self.mode = "Move"
            for i in self.scene().items():
                if hasattr(i, "ports"):
                    for p in i.ports.values():
                        p.setVisible(False)
            parent_widget = self.scene().views()[0].parentWidget()
            global_comps = getattr(
                parent_widget, "global_components", INDUSTRIAL_COMPONENTS
            )
            StreamEditorDialog(new_leg, global_components=global_comps).exec_()
        else:
            suggested = "Saída" if "right" in self.start_port.port_id else "Entrada"
            dialog = TerminalConfigDialog(self, suggested)
            if dialog.exec_() == QDialog.Accepted:
                v_type, v_name = dialog.get_values()
                if v_name:
                    handle = SourceSinkHandle(v_type, v_name)
                    handle.setPos(self.mapToScene(event.pos()))
                    self.scene().addItem(handle)

                    if v_type == "Entrada":
                        edge = Edge(
                            handle,
                            self.start_port.node,
                            "tip",
                            self.start_port.port_id,
                        )
                    else:
                        edge = Edge(
                            self.start_port.node,
                            handle,
                            self.start_port.port_id,
                            "tip",
                        )

                    edge.pipe_name = v_name
                    if v_type == "Saída":
                        edge.flow_data = self._get_outlet_flow_data(
                            self.start_port.node
                        )
                    else:
                        edge.flow_data = {}
                    self.scene().addItem(edge)
                    edge.adjust()
                    self.scene().removeItem(self.temp_line)
                    self.temp_line = None
                    self.mode = "Move"
                    for i in self.scene().items():
                        if hasattr(i, "ports"):
                            for p in i.ports.values():
                                p.setVisible(False)
                    parent_widget = self.scene().views()[0].parentWidget()
                    global_comps = getattr(
                        parent_widget, "global_components", INDUSTRIAL_COMPONENTS
                    )
                    StreamEditorDialog(edge, global_components=global_comps).exec_()
                    return
            self.scene().removeItem(self.temp_line)
            self.temp_line = None
            self.mode = "Move"
            for i in self.scene().items():
                if hasattr(i, "ports"):
                    for p in i.ports.values():
                        p.setVisible(False)

    def _get_outlet_flow_data(self, node):
        """
        Retorna flow_data para pré-preencher uma nova corrente de SAÍDA.
        Prioridade:
        1. stored_components do nó (resultado do balanço já calculado)
        2. Soma das correntes de entrada do nó (mistura das feeds)
        3. flow_data de correntes de saída existentes (se já calculadas)
        4. Dict vazio (fallback — usuário preenche manualmente)
        """
        stored = getattr(node, "stored_components", {})
        if stored and any(v > 0 for v in stored.values()):
            return {k: v for k, v in stored.items() if v > 0}

        in_edges = [e for e in node.edges if e.dest_node == node and e.flow_data]
        if in_edges:
            merged = {}
            for e in in_edges:
                for comp, val in e.flow_data.items():
                    merged[comp] = merged.get(comp, 0.0) + val
            if any(v > 0 for v in merged.values()):
                return merged

        out_edges = [
            e
            for e in node.edges
            if e.source_node == node
            and e.flow_data
            and any(v > 0 for v in e.flow_data.values())
        ]
        if out_edges:
            return out_edges[0].flow_data.copy()

        return {}

    def _probe_old_data(self, node):
        """Helper para recuperar flow_data em caso de reconexão."""
        return self._get_outlet_flow_data(node)

    def _get_input_flow_data(self, node):
        """Recupera os dados de fluxo das entradas de um nó para pré-preencher saídas."""
        return self._get_outlet_flow_data(node)


class FlowsheetWidget(QWidget):
    def __init__(self):
        super().__init__()
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)

        # Usa a toolbar padronizada do core (adicionando o botão de cálculo)
        self._tb, tb_lay = _make_toolbar(self, "PFD Flowsheet")
        self.btn_pal = QPushButton("☰ Equipamentos")
        self.btn_pal.setCheckable(True)
        self.btn_pal.setChecked(True)
        self.btn_solve = QPushButton("🧮 Calcular Balanço")
        self.btn_solve.clicked.connect(self.solve_mass_balance)

        self.btn_export = QPushButton("📊 Exportar Excel")
        self.btn_export.clicked.connect(self.export_to_excel)

        tb_lay.insertWidget(1, self.btn_pal)
        tb_lay.addWidget(self.btn_solve)
        tb_lay.addWidget(self.btn_export)
        layout.addWidget(self._tb)

        self.global_components = list(INDUSTRIAL_COMPONENTS)

        self.scene = QGraphicsScene()
        self.scene.setSceneRect(0, 0, 2000, 2000)
        self.canvas = FlowsheetCanvas(self.scene)

        # Splitter Horizontal (Paleta e Canvas)
        self.h_splitter = QSplitter(Qt.Horizontal)
        self.palette = SymbolPalette()
        self.h_splitter.addWidget(self.palette)
        self.h_splitter.addWidget(self.canvas)
        self.h_splitter.setStretchFactor(1, 1)
        self.h_splitter.setSizes([200, 800])

        # Tabela de Resultados do Balanço — Compacta e Elegante
        self.results_table = QTableWidget(self)
        self.results_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.results_table.setFixedSize(480, 280)  # Tamanho reduzido e proporcional
        self.results_table.setFrameShape(QTableWidget.NoFrame)
        self.results_table.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.results_table.setVerticalScrollMode(QTableWidget.ScrollPerPixel)
        self.results_table.setHorizontalScrollMode(QTableWidget.ScrollPerPixel)
        self.results_table.setShowGrid(True)
        self.results_table.setGridStyle(Qt.DotLine)
        self.results_table.raise_()

        layout.addWidget(self.h_splitter)

        self.process_name_input = QLineEdit(self.canvas)
        self.process_name_input.setPlaceholderText("Processo...")
        self.process_name_input.move(15, 15)

        # Timer para Autocalculação (Debounce de 300ms)
        self._auto_timer = QTimer(self)
        self._auto_timer.setSingleShot(True)
        self._auto_timer.timeout.connect(self.solve_mass_balance)
        self.scene.changed.connect(lambda: self._auto_timer.start(300))

        # Garante que as cores iniciais estejam corretas conforme o tema ativo
        self.refresh_theme()

    def solve_mass_balance(self):
        """
        Motor de Balanço de Massa — Felder & Rousseau Cap. 4 (Sistema Sem Reação, Estacionário)
        Implementa os 5 passos obrigatórios:
        1. Ordenação Topológica (BFS/Kahn)
        2. Identificar Feeds Conhecidas
        3. Para cada nó: Mistura → Distribuição → Verificação
        4. Verificação Global do Sistema
        5. Cálculo do Grau de Liberdade
        """
        edges = [i for i in self.scene.items() if isinstance(i, Edge)]
        nodes = [
            i for i in self.scene.items() if isinstance(i, (ProcessNode, JunctionNode))
        ]

        if not edges:
            return

        all_components = sorted(
            set(self.global_components + [c for e in edges for c in e.flow_data.keys()])
        )
        if not all_components:
            return

        components = all_components

        # Inicializa todas as correntes com zero para componentes ausentes
        for e in edges:
            for c in components:
                if c not in e.flow_data:
                    e.flow_data[c] = 0.0

        # Reseta atributos de todos os nós
        for node in nodes:
            node.stored_components = {c: 0.0 for c in components}
            node.total_input = {c: 0.0 for c in components}
            node.total_output = {c: 0.0 for c in components}
            node.mass_balance_error = 0.0
            node.has_steady_state = False

        # ============================================================
        # PASSO 1: ORDENAÇÃO TOPOLÓGICA (Algoritmo de Kahn)
        # ============================================================
        # Construir grafo dirigido: nodes -> nodes que dependem dele
        in_degree = {n: 0 for n in nodes}
        adjacency = {n: [] for n in nodes}  # n -> lista de nós que dependem de n

        for e in edges:
            if isinstance(e.source_node, (ProcessNode, JunctionNode)) and isinstance(
                e.dest_node, (ProcessNode, JunctionNode)
            ):
                src, dst = e.source_node, e.dest_node
                if src in nodes and dst in nodes:
                    adjacency[src].append(dst)
                    in_degree[dst] += 1

        # Detectando ciclos
        queue = [n for n in nodes if in_degree[n] == 0]
        topo_order = []
        while queue:
            node = queue.pop(0)
            topo_order.append(node)
            for neighbor in adjacency[node]:
                in_degree[neighbor] -= 1
                if in_degree[neighbor] == 0:
                    queue.append(neighbor)

        if len(topo_order) != len(nodes):
            QMessageBox.warning(
                self,
                "Ciclo Detectado",
                "⚠️ O sistema contém ciclos (reciclo) que não podem ser resolvidos "
                "por ordenação topológica simples.\n\n"
                "Para reciclos, use o modo iterativo (a ser implementado).",
            )
            return

        # ============================================================
        # PASSO 2: IDENTIFICAR FEEDS CONHECIDAS
        # ============================================================
        feed_edges = [
            e
            for e in edges
            if isinstance(e.source_node, SourceSinkHandle)
            and e.source_node.h_type == "Entrada"
            and any(v > 0 for v in e.flow_data.values())
        ]
        feed_set = set(feed_edges)

        # ============================================================
        # PASSO 3: PROCESSAR CADA NÓ NA ORDEM TOPOLÓGICA
        # ============================================================
        for node in topo_order:
            in_edges = [e for e in node.edges if e.dest_node == node]
            out_edges = [e for e in node.edges if e.source_node == node]

            if not in_edges:
                continue  # Nó sem entradas (só fontes externas)

            # === 3a. MISTURA: somar todas as entradas ===
            total_in = {c: 0.0 for c in components}
            for e in in_edges:
                for c in components:
                    total_in[c] += e.flow_data.get(c, 0.0)

            node.total_input = total_in.copy()
            node.stored_components = total_in.copy()

            if not out_edges:
                continue  # Nó sem saídas (sumidouro)

            # === 3b. DISTRIBUIÇÃO nas saídas (usando split_config) ===
            split_config = getattr(node, "split_config", {})

            for comp in components:
                if total_in[comp] < 1e-9:
                    # Sem entrada deste componente: distribui zero
                    for oe in out_edges:
                        oe.flow_data[comp] = 0.0
                    continue

                # Separa saídas por modo
                fixed_outputs = []
                perc_outputs = []

                comp_config = split_config.get(comp, {})
                for oe in out_edges:
                    port = oe.source_port
                    cfg = comp_config.get(
                        port, {"mode": "perc", "val": 1.0 / len(out_edges)}
                    )
                    if cfg["mode"] == "fixed":
                        fixed_outputs.append((oe, cfg["val"]))
                    else:
                        perc_outputs.append((oe, cfg["val"]))

                remaining = total_in[comp]

                # 1º: Processa saídas de vazão FIXA
                for oe, fixed_val in fixed_outputs:
                    val = min(remaining, fixed_val)
                    oe.flow_data[comp] = val
                    remaining -= val

                # 2º: Distribui restante por FRAÇÃO
                if perc_outputs and remaining > 1e-9:
                    total_perc = sum(p[1] for p in perc_outputs)
                    if total_perc > 0:
                        # Normaliza para soma = 1
                        for oe, orig_val in perc_outputs:
                            share = orig_val / total_perc
                            oe.flow_data[comp] = remaining * share
                    else:
                        # Sem config: divide igualmente
                        share = remaining / len(perc_outputs)
                        for oe, _ in perc_outputs:
                            oe.flow_data[comp] = share
                elif remaining > 1e-9 and out_edges:
                    # Sem configuração nenhuma: distribui igualmente
                    share = remaining / len(out_edges)
                    for oe in out_edges:
                        oe.flow_data[comp] = share

                # Garante componentes ausentes como zero
                for oe in out_edges:
                    if comp not in oe.flow_data:
                        oe.flow_data[comp] = 0.0

            # Atualiza totais de saída do nó
            for oe in out_edges:
                for c in components:
                    node.total_output[c] = node.total_output.get(
                        c, 0.0
                    ) + oe.flow_data.get(c, 0.0)

            # === 3c. VERIFICAÇÃO de conservação por nó ===
            total_out = {c: 0.0 for c in components}
            for oe in out_edges:
                for c in components:
                    total_out[c] += oe.flow_data.get(c, 0.0)

            node_error = 0.0
            for c in components:
                err = abs(total_in[c] - total_out[c])
                node_error += err
            node.mass_balance_error = node_error
            node.has_steady_state = node_error < 1e-6

        # ============================================================
        # PASSO 4: VERIFICAÇÃO GLOBAL DO SISTEMA
        # ============================================================
        total_feed = sum(
            sum(e.flow_data.values())
            for e in edges
            if isinstance(e.source_node, SourceSinkHandle)
            and e.source_node.h_type == "Entrada"
        )
        total_product = sum(
            sum(e.flow_data.values())
            for e in edges
            if isinstance(e.dest_node, SourceSinkHandle)
            and e.dest_node.h_type == "Saída"
        )

        # Massa retida nos equipamentos
        total_stored = 0.0
        for node in nodes:
            stored = getattr(node, "stored_components", {})
            total_stored += sum(stored.values())

        global_error = abs(total_feed - total_product - total_stored)
        closure = (global_error / total_feed * 100) if total_feed > 1e-9 else 0.0

        # === Grau de Liberdade ===
        gl, n_vars, n_eqs, unknowns, _ = self._compute_degrees_of_freedom()

        if gl > 0:
            msg = (
                f"⚠️ Sistema sub-especificado (GL = +{gl}).\n"
                f"Variáveis: {n_vars}  |  Equações: {n_eqs}\n\n"
                f"Forneça mais dados nas correntes de entrada."
            )
            QMessageBox.warning(self, "Grau de Liberdade", msg)

        # === Relatório de Resultados ===
        nodes_with_error = [
            (n, n.mass_balance_error) for n in nodes if n.mass_balance_error > 1e-6
        ]
        significant_nodes = [
            (n, e) for n, e in nodes_with_error if e > 0.001 * max(total_feed, 1.0)
        ]

        if significant_nodes:
            error_msg = "⚠️ Erros de conservação detectados nos equipamentos:\n\n"
            for node, err in significant_nodes[:5]:
                name = getattr(node, "custom_name", "") or getattr(
                    node, "symbol_type", ""
                )
                if not name:
                    name = "Equipamento"
                stored = getattr(node, "stored_components", {})
                total_stored_node = sum(stored.values())
                if total_stored_node > 0.001:
                    error_msg += f"• {name}: desvio {err:.3f} kg/h (retido: {total_stored_node:.2f})\n"
                else:
                    error_msg += f"• {name}: desvio {err:.3f} kg/h\n"
            if len(significant_nodes) > 5:
                error_msg += f"... e mais {len(significant_nodes) - 5} equipamentos.\n"
            error_msg += "\nVerifique as conexões e configurações de split."
            QMessageBox.warning(self, "Conservação de Massa", error_msg)
        elif closure > 0.01 and total_stored < 0.001:
            stored_msg = (
                f"Massa retida: {total_stored:.2f} kg/h\n"
                if total_stored > 0.001
                else ""
            )
            QMessageBox.warning(
                self,
                "Balanço Global",
                f"⚠️ Erro de fechamento: {closure:.4f}%\n"
                f"Entrada: {total_feed:.2f} kg/h\n"
                f"Saída:   {total_product:.2f} kg/h\n"
                f"{stored_msg}Verifique se todas as saídas estão conectadas.",
            )

        # Ajusta todas as arestas
        for e in edges:
            e.adjust()

        # Atualiza tabela de resultados
        self.update_results_table(edges)

    def update_results_table(self, _=None):
        """
        Atualiza a tabela de resultados com os valores do balanço de massa.
        Exibe todas as correntes (edges) e equipamentos (nodes) com seus fluxos.
        """
        # Força a atualização da cena para garantir dados atualizados
        self.scene.update()

        # Coleta todos os itens
        all_items = self.scene.items()
        edges = [i for i in all_items if isinstance(i, Edge)]
        nodes = [
            i
            for i in all_items
            if isinstance(i, (ProcessNode, JunctionNode, SourceSinkHandle))
        ]

        # Coleta componentes únicos de TODAS as correntes (inclui global_components)
        components = sorted(
            set(self.global_components + [c for e in edges for c in e.flow_data.keys()])
        )
        if not components:
            components = ["Fluxo"]  # Coluna padrão se não houver componentes definidos

        # Headers: Tipo, Nome, Origem → Destino, Entrada, Saída, Total, Retido, Componentes...
        headers = [
            "Tipo",
            "Nome",
            "Rota",
            "Entrada",
            "Saída",
            "Total",
            "Retido",
        ] + components
        self.results_table.setColumnCount(len(headers))
        self.results_table.setHorizontalHeaderLabels(headers)

        # Calcula GL para tooltip
        gl, n_vars, n_eqs, _, _ = self._compute_degrees_of_freedom()
        gl_text = f"GL={gl}" if gl != 0 else "GL=0 ✅"
        self.results_table.setToolTip(
            f"Balanço de Massa | {gl_text} | Variáveis: {n_vars} | Equações: {n_eqs}"
        )

        # Lista de entradas: Correntes primeiro, depois equipamentos
        all_entries = edges + nodes
        self.results_table.setRowCount(len(all_entries))

        row_idx = 0
        # ==================== CORRENTES (EDGES) ====================
        for edge in edges:
            # Tipo
            tipo_item = QTableWidgetItem("🌊 Corrente")
            tipo_item.setTextAlignment(Qt.AlignCenter)

            # Nome
            name = edge.pipe_name if edge.pipe_name else f"C{row_idx + 1}"
            name_item = QTableWidgetItem(name)
            name_item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)

            # Rota: Origem → Destino
            if isinstance(edge.source_node, ProcessNode):
                src = edge.source_node.custom_name or edge.source_node.symbol_type[:8]
            elif isinstance(edge.source_node, SourceSinkHandle):
                src = edge.source_node.flow_name or f"{edge.source_node.h_type[:3]}"
            else:
                src = "?"

            if isinstance(edge.dest_node, ProcessNode):
                dst = edge.dest_node.custom_name or edge.dest_node.symbol_type[:8]
            elif isinstance(edge.dest_node, SourceSinkHandle):
                dst = edge.dest_node.flow_name or f"{edge.dest_node.h_type[:3]}"
            else:
                dst = "?"

            rota_item = QTableWidgetItem(f"{src} → {dst}")
            rota_item.setTextAlignment(Qt.AlignCenter)
            rota_item.setToolTip(f"Origem: {src}\nDestino: {dst}")

            # Valores de fluxo
            total_flow = sum(edge.flow_data.values())
            entrada_item = QTableWidgetItem("-")
            entrada_item.setTextAlignment(Qt.AlignCenter)

            saida_item = QTableWidgetItem("-")
            saida_item.setTextAlignment(Qt.AlignCenter)

            total_item = QTableWidgetItem(f"{total_flow:.2f}")
            total_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            total_item.setFont(QFont("Consolas", 9, QFont.Bold))

            # Destaca correntes com fluxo zero
            if total_flow < 0.001:
                total_item.setForeground(QBrush(QColor("#888888")))
            else:
                total_item.setForeground(QBrush(QColor(T()["accent_bright"])))

            # Preenche células
            self.results_table.setItem(row_idx, 0, tipo_item)
            self.results_table.setItem(row_idx, 1, name_item)
            self.results_table.setItem(row_idx, 2, rota_item)
            self.results_table.setItem(row_idx, 3, entrada_item)
            self.results_table.setItem(row_idx, 4, saida_item)
            self.results_table.setItem(row_idx, 5, total_item)
            retido_item = QTableWidgetItem("-")
            retido_item.setTextAlignment(Qt.AlignCenter)
            self.results_table.setItem(row_idx, 6, retido_item)

            # Componentes
            for j, comp in enumerate(components):
                val = edge.flow_data.get(comp, 0.0)
                comp_item = QTableWidgetItem(f"{val:.2f}")
                comp_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                comp_item.setFont(QFont("Consolas", 8))
                if val < 0.001:
                    comp_item.setForeground(QBrush(QColor("#666666")))
                self.results_table.setItem(row_idx, 7 + j, comp_item)

            row_idx += 1

        # ==================== EQUIPAMENTOS (NODES) ====================
        for node in nodes:
            in_edges = [e for e in edges if e.dest_node == node]
            out_edges = [e for e in edges if e.source_node == node]

            mass_in = sum(sum(e.flow_data.values()) for e in in_edges)
            mass_out = sum(sum(e.flow_data.values()) for e in out_edges)

            # Tipo e nome
            if isinstance(node, ProcessNode):
                tipo = f"⚙️ {node.symbol_type[:10]}"
                name = node.custom_name or node.symbol_type[:12]
            elif isinstance(node, JunctionNode):
                tipo = "💠 Junção"
                name = "T-junction"
            else:  # SourceSinkHandle
                tipo = f"🚉 {node.h_type}"
                name = node.flow_name or "Terminal"

            tipo_item = QTableWidgetItem(tipo)
            tipo_item.setTextAlignment(Qt.AlignCenter)

            name_item = QTableWidgetItem(name)
            name_item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)

            # Rota: resume conexões
            if in_edges:
                srcs = [e.pipe_name or "?" for e in in_edges[:2]]
                rota_in = (
                    f"{len(in_edges)}E" if len(in_edges) > 2 else f"{'+'.join(srcs)}"
                )
            else:
                rota_in = "—"

            if out_edges:
                dsts = [e.pipe_name or "?" for e in out_edges[:2]]
                rota_out = (
                    f"{len(out_edges)}S" if len(out_edges) > 2 else f"{'+'.join(dsts)}"
                )
            else:
                rota_out = "—"

            rota_item = QTableWidgetItem(f"{rota_in}→{rota_out}")
            rota_item.setTextAlignment(Qt.AlignCenter)
            rota_item.setToolTip(
                f"Entradas: {len(in_edges)} | Saídas: {len(out_edges)}"
            )

            # Entrada e Saída do equipamento
            entrada_item = QTableWidgetItem(f"{mass_in:.2f}" if mass_in > 0 else "-")
            entrada_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            entrada_item.setFont(QFont("Consolas", 8))

            saida_item = QTableWidgetItem(f"{mass_out:.2f}" if mass_out > 0 else "-")
            saida_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            saida_item.setFont(QFont("Consolas", 8))

            # Total = entrada (ou saída se não houver entrada)
            total = mass_in if mass_in > 0 else mass_out
            total_item = QTableWidgetItem(f"{total:.2f}")
            total_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            total_item.setFont(QFont("Consolas", 9, QFont.Bold))

            # Indicador de balanço no equipamento
            if in_edges and out_edges:
                diff = abs(mass_in - mass_out)
                if diff < 0.001:
                    total_item.setBackground(QBrush(QColor("#1a472a")))  # Verde escuro
                    total_item.setForeground(QBrush(QColor("#4ade80")))  # Verde claro
                elif diff < 0.1 * max(mass_in, 1):
                    total_item.setBackground(
                        QBrush(QColor("#854d0e"))
                    )  # Amarelo escuro
                    total_item.setForeground(QBrush(QColor("#facc15")))  # Amarelo
                else:
                    total_item.setBackground(
                        QBrush(QColor("#7f1d1d"))
                    )  # Vermelho escuro
                    total_item.setForeground(
                        QBrush(QColor("#f87171"))
                    )  # Vermelho claro

            # Preenche células
            self.results_table.setItem(row_idx, 0, tipo_item)
            self.results_table.setItem(row_idx, 1, name_item)
            self.results_table.setItem(row_idx, 2, rota_item)
            self.results_table.setItem(row_idx, 3, entrada_item)
            self.results_table.setItem(row_idx, 4, saida_item)
            self.results_table.setItem(row_idx, 5, total_item)

            stored = getattr(node, "stored_components", {})
            total_stored = sum(stored.values())
            retido_item = QTableWidgetItem(
                f"{total_stored:.2f}" if total_stored > 0 else "-"
            )
            retido_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            retido_item.setFont(QFont("Consolas", 8))
            if total_stored > 0.001:
                retido_item.setForeground(
                    QBrush(QColor("#f97316"))
                )  # Orange for retained
            self.results_table.setItem(row_idx, 6, retido_item)

            # Componentes (soma das entradas)
            comps_data = {
                c: sum(e.flow_data.get(c, 0) for e in in_edges) for c in components
            }
            for j, comp in enumerate(components):
                val = comps_data.get(comp, 0.0)
                comp_item = QTableWidgetItem(f"{val:.2f}")
                comp_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                comp_item.setFont(QFont("Consolas", 8))
                if val < 0.001:
                    comp_item.setForeground(QBrush(QColor("#666666")))
                self.results_table.setItem(row_idx, 6 + j, comp_item)

            row_idx += 1

        # Ajuste de colunas
        self.results_table.horizontalHeader().setSectionResizeMode(
            QHeaderView.Interactive
        )
        self.results_table.horizontalHeader().setSectionResizeMode(
            0, QHeaderView.ResizeToContents
        )  # Tipo
        self.results_table.horizontalHeader().setSectionResizeMode(
            1, QHeaderView.Stretch
        )  # Nome
        self.results_table.horizontalHeader().setSectionResizeMode(
            2, QHeaderView.ResizeToContents
        )  # Rota
        self.results_table.horizontalHeader().setSectionResizeMode(
            3, QHeaderView.ResizeToContents
        )  # Entrada
        self.results_table.horizontalHeader().setSectionResizeMode(
            4, QHeaderView.ResizeToContents
        )  # Saída
        self.results_table.horizontalHeader().setSectionResizeMode(
            5, QHeaderView.ResizeToContents
        )  # Total
        self.results_table.horizontalHeader().setSectionResizeMode(
            6, QHeaderView.ResizeToContents
        )  # Retido

        # Ajuste de largura mínima para colunas de componentes
        for j in range(len(components)):
            self.results_table.setColumnWidth(7 + j, 60)

    def _compute_degrees_of_freedom(self):
        """
        Calcula o Grau de Liberdade seguindo Felder Cap. 4.
        Retorna: (gl: int, n_vars: int, n_eqs: int, unknowns: list[Edge], components: set)
        """
        edges = [i for i in self.scene.items() if isinstance(i, Edge)]
        nodes = [
            i for i in self.scene.items() if isinstance(i, (ProcessNode, JunctionNode))
        ]

        # Identifica todos os componentes presentes no processo
        components = set()
        for e in edges:
            components.update(e.flow_data.keys())

        # Correntes desconhecidas: correntes internas (não-feed) com flow_data vazio ou zerado
        feed_edges = {
            e
            for e in edges
            if isinstance(e.source_node, SourceSinkHandle)
            and e.source_node.h_type == "Entrada"
            and any(v > 0 for v in e.flow_data.values())
        }

        unknowns = [e for e in edges if e not in feed_edges]

        # Número de incógnitas: cada corrente desconhecida × número de componentes
        n_vars = len(unknowns) * max(len(components), 1)

        # Número de equações: por nó, uma equação por componente
        n_eqs = len(nodes) * max(len(components), 1)

        gl = n_vars - n_eqs
        return gl, n_vars, n_eqs, unknowns, components

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self.update_table_position()

    def update_table_position(self):
        ww = self.width()
        wh = self.height()
        tw = self.results_table.width()
        th = self.results_table.height()
        # Posiciona no canto inferior direito com padding reduzido de 15px
        self.results_table.move(ww - tw - 15, wh - th - 15)
        self.results_table.raise_()

    def refresh_theme(self):
        t = T()
        # Estiliza o botão "Equipamentos" com tipografia controlada para não "estourar"
        self.btn_pal.setStyleSheet(f"""
            QPushButton {{ 
                background: transparent; color: {t["text"]}; 
                font-weight: bold; border: none; padding: 2px 10px; 
                font-size: 13px; font-family: 'Segoe UI';
            }}
            QPushButton:hover {{ color: {t["accent"]}; }}
            QPushButton:checked {{ color: {t["accent_bright"]}; text-decoration: underline; }}
        """)
        self.btn_solve.setStyleSheet(f"""
            QPushButton {{ 
                background: {t["accent"]}; color: white; 
                font-weight: bold; border-radius: 4px; padding: 5px 15px; 
            }}
            QPushButton:hover {{ background: {t["accent_bright"]}; }}
        """)
        self.btn_export.setStyleSheet(f"""
            QPushButton {{ 
                background: {t["accent"]}; color: white; 
                font-weight: bold; border-radius: 4px; padding: 5px 15px; 
            }}
            QPushButton:hover {{ background: {t["accent_bright"]}; }}
        """)
        self._tb.setStyleSheet(
            f"background:{t['toolbar_bg']}; border-bottom:2px solid {t['accent']};"
        )
        self.canvas.setBackgroundBrush(QBrush(QColor(t["bg_app"])))

        # Estilo Glassmorphism Compacto e Uniforme para a Tabela
        rgba = QColor(t["bg_card"])
        glass_bg = f"rgba({rgba.red()}, {rgba.green()}, {rgba.blue()}, 200)"
        glass_border = f"rgba({rgba.red()}, {rgba.green()}, {rgba.blue()}, 255)"

        self.results_table.setStyleSheet(f"""
            QTableWidget {{
                background: {glass_bg};
                color: {t["text"]};
                gridline-color: {t["accent_dim"]};
                border: 1px solid {t["accent"]};
                border-radius: 8px;
                font-family: 'Segoe UI', Arial, sans-serif;
                font-size: 9px;
                outline: none;
            }}
            QTableWidget::item {{
                padding: 4px 6px;
                border-bottom: 1px solid {t["accent_dim"]};
                border-right: 1px solid {t["accent_dim"]};
            }}
            QTableWidget::item:selected {{
                background-color: {t["accent_dim"]};
                color: {t["text"]};
            }}
            QTableWidget::item:hover {{
                background-color: {t["bg_card2"]};
            }}
            QHeaderView::section {{
                background-color: {t["accent"]};
                color: white;
                border: none;
                border-right: 1px solid {t["accent_dim"]};
                border-bottom: 1px solid {t["accent_dim"]};
                font-weight: 600;
                padding: 5px 4px;
                font-size: 8px;
                text-transform: uppercase;
                letter-spacing: 0.3px;
            }}
            QHeaderView::section:last {{
                border-right: none;
            }}
            QHeaderView::section:horizontal {{
                border-top-left-radius: 0px;
                border-top-right-radius: 0px;
            }}
            QScrollBar:vertical {{
                border: none;
                background: transparent;
                width: 6px;
                margin: 2px;
            }}
            QScrollBar::handle:vertical {{
                background: {t["accent_dim"]};
                min-height: 20px;
                border-radius: 3px;
            }}
            QScrollBar::handle:vertical:hover {{
                background: {t["accent"]};
            }}
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
                height: 0px;
            }}
            QScrollBar:horizontal {{
                border: none;
                background: transparent;
                height: 6px;
                margin: 2px;
            }}
            QScrollBar::handle:horizontal {{
                background: {t["accent_dim"]};
                min-width: 20px;
                border-radius: 3px;
            }}
            QScrollBar::handle:horizontal:hover {{
                background: {t["accent"]};
            }}
            QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {{
                width: 0px;
            }}
        """)

        # Força o reposicionamento inicial
        self.update_table_position()
        if hasattr(self, "palette"):
            self.palette.refresh_theme()
        # Força redo em todos os itens da cena para pegar novas cores e máscaras
        for item in self.scene.items():
            if isinstance(item, Edge):
                item.adjust()  # Recalcula máscara com nova cor bg_app
            item.update()

    def export_to_excel(self):
        """Exporta os dados de balanço de massa para Excel usando openpyxl."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            QMessageBox.warning(
                self,
                "Erro",
                "Biblioteca openpyxl não está instalada.\nExecute: pip install openpyxl",
            )
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Salvar Balanço de Massa Excel",
            f"{self.process_name_input.text() or 'balanco'}_massa.xlsx",
            "Excel Files (*.xlsx)",
        )
        if not file_path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Balanço de Massa"

        all_items = self.scene.items()
        edges = [i for i in all_items if isinstance(i, Edge)]
        nodes = [
            i
            for i in all_items
            if isinstance(i, (ProcessNode, JunctionNode, SourceSinkHandle))
        ]

        components = sorted(
            set(self.global_components + [c for e in edges for c in e.flow_data.keys()])
        )
        if not components:
            components = ["Fluxo"]

        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        headers = [
            "Tipo",
            "Nome",
            "Rota",
            "Entrada (kg/h)",
            "Saída (kg/h)",
            "Total (kg/h)",
            "Retido (kg/h)",
        ] + components
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin_border

        row = 2
        for edge in edges:
            if isinstance(edge.source_node, ProcessNode):
                src = edge.source_node.custom_name or edge.source_node.symbol_type
            elif isinstance(edge.source_node, SourceSinkHandle):
                src = edge.source_node.flow_name or edge.source_node.h_type
            else:
                src = "?"

            if isinstance(edge.dest_node, ProcessNode):
                dst = edge.dest_node.custom_name or edge.dest_node.symbol_type
            elif isinstance(edge.dest_node, SourceSinkHandle):
                dst = edge.dest_node.flow_name or edge.dest_node.h_type
            else:
                dst = "?"

            total_flow = sum(edge.flow_data.values())

            data = [
                "Corrente",
                edge.pipe_name or f"C{row - 1}",
                f"{src} → {dst}",
                "-",
                "-",
                f"{total_flow:.2f}",
                "-",
            ]
            for comp in components:
                data.append(f"{edge.flow_data.get(comp, 0.0):.2f}")

            for col, val in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")
            row += 1

        for node in nodes:
            in_edges = [e for e in edges if e.dest_node == node]
            out_edges = [e for e in edges if e.source_node == node]

            mass_in = sum(sum(e.flow_data.values()) for e in in_edges)
            mass_out = sum(sum(e.flow_data.values()) for e in out_edges)
            stored = getattr(node, "stored_components", {})
            total_stored = sum(stored.values())

            if isinstance(node, ProcessNode):
                tipo = node.symbol_type
                nome = node.custom_name or node.symbol_type
            elif isinstance(node, JunctionNode):
                tipo = "Junção"
                nome = "T-junction"
            else:
                tipo = node.h_type
                nome = node.flow_name or "Terminal"

            srcs = [e.pipe_name or "?" for e in in_edges[:2]]
            dsts = [e.pipe_name or "?" for e in out_edges[:2]]
            rota = (
                f"{'+'.join(srcs) if srcs else '—'}→{'+'.join(dsts) if dsts else '—'}"
            )

            total = mass_in if mass_in > 0 else mass_out

            data = [
                tipo,
                nome,
                rota,
                f"{mass_in:.2f}" if mass_in > 0 else "-",
                f"{mass_out:.2f}" if mass_out > 0 else "-",
                f"{total:.2f}",
                f"{total_stored:.2f}" if total_stored > 0 else "-",
            ]
            for comp in components:
                val = sum(e.flow_data.get(comp, 0) for e in in_edges)
                data.append(f"{val:.2f}")

            for col, val in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")
            row += 1

        for col in range(1, len(headers) + 1):
            if col <= 7:
                ws.column_dimensions[get_column_letter(col)].width = 18
            else:
                ws.column_dimensions[get_column_letter(col)].width = 12

        try:
            wb.save(file_path)
            QMessageBox.information(self, "Sucesso", f"Arquivo salvo:\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Não foi possível salvar:\n{str(e)}")

    def load_example(self, example_id):
        # 1. Limpar cena
        self.scene.clear()
        self.process_name_input.setText(example_id)

        # 2. Dados dos Exemplos
        examples = {
            "Refinaria": {
                "nodes": [
                    ("Vaso Horizontal", 50, 350, "Carga"),
                    ("Bomba Centrífuga", 180, 360, "P-101"),
                    ("Torre de Destilação", 350, 250, "T-101"),
                    ("Condensador de Casco", 550, 150, "E-101"),
                    ("Tanque Fechado", 750, 150, "Destilado"),
                ],
                "edges": [
                    (0, 1, "right", "left"),
                    (1, 2, "right", "left"),
                    (2, 3, "top", "left"),
                    (3, 4, "right", "left"),
                ],
            },
            "Produção de Amônia": {
                "nodes": [
                    ("Compressor", 80, 300, "C-101"),
                    ("Reator Tubular (PFR)", 300, 300, "R-101"),
                    ("Trocador de Placas", 530, 300, "E-102"),
                    ("Separador Bifásico", 750, 300, "V-102"),
                ],
                "edges": [
                    (0, 1, "right", "left"),
                    (1, 2, "right", "left"),
                    (2, 3, "right", "left"),
                ],
            },
            "Tratamento de Água (ETA)": {
                "nodes": [
                    ("Clarificador", 100, 250, "Clarificador"),
                    ("Filtro de Mangas", 300, 250, "F-101"),
                    ("Filtro Cartucho", 500, 250, "F-102"),
                    ("Tanque Aberto", 700, 250, "Água Tratada"),
                ],
                "edges": [
                    (0, 1, "right", "left"),
                    (1, 2, "right", "left"),
                    (2, 3, "right", "left"),
                ],
            },
            "Caldeira Industrial": {
                "nodes": [
                    ("Tanque Aberto", 80, 450, "Água Alimentação"),
                    ("Bomba Centrífuga", 220, 460, "B-201"),
                    ("Caldeira", 420, 380, "Vaporizador"),
                    ("Chaminé", 650, 200, "Exaustão"),
                ],
                "edges": [
                    (0, 1, "right", "left"),
                    (1, 2, "right", "left"),
                    (2, 3, "top", "bottom"),
                ],
            },
            "Linha de Mineração": {
                "nodes": [
                    ("Britador de Mandíbula", 100, 200, "Britagem"),
                    ("Correia Transportadora", 350, 220, "Transporte"),
                    ("Moinho", 600, 200, "Moagem"),
                ],
                "edges": [(0, 1, "right", "left"), (1, 2, "right", "left")],
            },
        }

        if example_id not in examples:
            return
        ex = examples[example_id]

        # 3. Criar Nodes
        created_nodes = []
        for sym, x, y, name in ex["nodes"]:
            node = ProcessNode(sym)
            node.setPos(x, y)
            node.set_name(name)
            self.scene.addItem(node)
            created_nodes.append(node)

        # 4. Criar Edges
        for s_idx, t_idx, s_por, t_por in ex["edges"]:
            edge = Edge(created_nodes[s_idx], created_nodes[t_idx], s_por, t_por)
            self.scene.addItem(edge)
            edge.adjust()

    def refresh_theme(self):
        t = T()
        # Estiliza o botão "Equipamentos" com tipografia controlada para não "estourar"
        self.btn_pal.setStyleSheet(f"""
            QPushButton {{ 
                background: transparent; color: {t["text"]}; 
                font-weight: bold; border: none; padding: 2px 10px; 
                font-size: 13px; font-family: 'Segoe UI';
            }}
            QPushButton:hover {{ color: {t["accent"]}; }}
            QPushButton:checked {{ color: {t["accent_bright"]}; text-decoration: underline; }}
        """)
        self._tb.setStyleSheet(
            f"background:{t['toolbar_bg']}; border-bottom:2px solid {t['accent']};"
        )
        self.canvas.setBackgroundBrush(QBrush(QColor(t["bg_app"])))
        if hasattr(self, "palette"):
            self.palette.refresh_theme()
        # Força redo em todos os itens da cena para pegar novas cores e máscaras
        for item in self.scene.items():
            if isinstance(item, Edge):
                item.adjust()  # Recalcula máscara com nova cor bg_app
            item.update()


class _FlowsheetModule(BaseModule):
    def __init__(self):
        super().__init__()
        self._inner = FlowsheetWidget()
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.addWidget(self._inner)

    # --- BaseModule API -------------------------------------------------
    def get_state(self):
        scene_items = self._inner.scene.items()
        nodes = []
        handles = []
        edges = []
        node_ids = {}
        handle_ids = {}

        # 1) Primeiro coletamos os nós e terminais para criar mapas de ID
        for i in scene_items:
            if isinstance(i, ProcessNode):
                node_id = f"n{len(node_ids) + 1}"
                node_ids[i] = node_id
                nodes.append(
                    {
                        "id": node_id,
                        "type": i.symbol_type,
                        "x": i.scenePos().x(),
                        "y": i.scenePos().y(),
                        "name": getattr(i, "custom_name", ""),
                        "size": getattr(i, "size", 50),
                        "split_config": getattr(i, "split_config", {}),
                        "stored_components": getattr(i, "stored_components", {}),
                        "total_input": getattr(i, "total_input", {}),
                        "total_output": getattr(i, "total_output", {}),
                        "mass_balance_error": getattr(i, "mass_balance_error", 0.0),
                        "has_steady_state": getattr(i, "has_steady_state", False),
                    }
                )
            elif isinstance(i, SourceSinkHandle):
                handle_id = f"h{len(handle_ids) + 1}"
                handle_ids[i] = handle_id
                handles.append(
                    {
                        "id": handle_id,
                        "h_type": i.h_type,
                        "flow_name": i.flow_name,
                        "x": i.scenePos().x(),
                        "y": i.scenePos().y(),
                    }
                )

        # 2) Depois coletamos conexões usando IDs estáveis
        for i in scene_items:
            if not isinstance(i, Edge):
                continue

            source_kind = "node" if isinstance(i.source_node, ProcessNode) else "handle"
            dest_kind = "node" if isinstance(i.dest_node, ProcessNode) else "handle"
            source_ref = (
                node_ids.get(i.source_node)
                if source_kind == "node"
                else handle_ids.get(i.source_node)
            )
            dest_ref = (
                node_ids.get(i.dest_node)
                if dest_kind == "node"
                else handle_ids.get(i.dest_node)
            )

            # Se algum endpoint não está mapeado, ignora esta aresta para evitar estado inconsistente
            if not source_ref or not dest_ref:
                continue

            edges.append(
                {
                    "source_kind": source_kind,
                    "source_id": source_ref,
                    "source_port": i.source_port,
                    "dest_kind": dest_kind,
                    "dest_id": dest_ref,
                    "dest_port": i.dest_port,
                    "pipe_name": i.pipe_name,
                    "is_utility": i.is_utility,
                    "flow_data": i.flow_data,
                }
            )

        return {
            "schema": "flowsheet.v6",
            "process_name": self._inner.process_name_input.text(),
            "nodes": nodes,
            "handles": handles,
            "edges": edges,
            "global_components": getattr(
                self._inner, "global_components", INDUSTRIAL_COMPONENTS
            ),
        }

    def set_state(self, state):
        self._inner.scene.clear()

        # Compatibilidade com estado antigo (v1): {"nodes": [{"type","x","y"}]}
        if state.get("schema") != "flowsheet.v2":
            for n in state.get("nodes", []):
                node = ProcessNode(n["type"])
                node.setPos(n["x"], n["y"])
                self._inner.scene.addItem(node)
            self._inner.process_name_input.setText(state.get("process_name", ""))
            return

        created_nodes = {}
        created_handles = {}

        # 1) Recria nós de processo
        for n in state.get("nodes", []):
            node = ProcessNode(n.get("type", ""))
            node.setPos(n.get("x", 0), n.get("y", 0))
            if n.get("name"):
                node.set_name(n.get("name", ""))
            if "size" in n:
                try:
                    node.set_size(int(n["size"]))
                except Exception:
                    pass
            if "split_config" in n:
                node.split_config = n["split_config"]
            # Restaura dados de armazenamento
            if "stored_components" in n:
                node.stored_components = n["stored_components"]
            if "total_input" in n:
                node.total_input = n["total_input"]
            if "total_output" in n:
                node.total_output = n["total_output"]
            if "mass_balance_error" in n:
                node.mass_balance_error = n["mass_balance_error"]
            if "has_steady_state" in n:
                node.has_steady_state = n["has_steady_state"]
            self._inner.scene.addItem(node)
            created_nodes[n.get("id")] = node

        # 2) Recria terminais source/sink
        for h in state.get("handles", []):
            handle = SourceSinkHandle(h.get("h_type", "Saída"), h.get("flow_name", ""))
            handle.setPos(h.get("x", 0), h.get("y", 0))
            self._inner.scene.addItem(handle)
            created_handles[h.get("id")] = handle

        # 3) Recria conexões
        for e in state.get("edges", []):
            src = (
                created_nodes.get(e.get("source_id"))
                if e.get("source_kind") == "node"
                else created_handles.get(e.get("source_id"))
            )
            dst = (
                created_nodes.get(e.get("dest_id"))
                if e.get("dest_kind") == "node"
                else created_handles.get(e.get("dest_id"))
            )
            if src is None or dst is None:
                continue
            edge = Edge(
                src, dst, e.get("source_port", "right"), e.get("dest_port", "left")
            )
            edge.pipe_name = e.get("pipe_name", "")
            edge.is_utility = bool(e.get("is_utility", False))
            edge.flow_data = e.get("flow_data", {})
            edge.adjust()
            self._inner.scene.addItem(edge)

        self._inner.process_name_input.setText(state.get("process_name", ""))

        if "global_components" in state:
            self._inner.global_components = state["global_components"]

    def refresh_theme(self):
        if hasattr(self._inner, "refresh_theme"):
            self._inner.refresh_theme()

    def get_view(self):
        return getattr(self._inner, "canvas", None)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    w = _FlowsheetModule()
    w.show()
    sys.exit(app.exec_())
